"""Excel batch run, navigation, claim extraction, and row pipeline."""
import os
import re
import tempfile
import threading
import time
import traceback
from tkinter import messagebox

import pandas as pd
from playwright.sync_api import expect

from availity_app import state
from availity_app.constants import (
    AVAILITY_NAVIGATION_HOME,
    CLAIM_NOT_FOUND_STATUS,
    DOB_NOT_FOUND_STATUS,
    INVALID_MEMBER_ID_STATUS,
    KNOWN_ERROR_STATUSES,
    NPI,
    PAID_OVERPAYMENT_STATUS,
    PAYER_SEARCH_ERROR_STATUS,
    OUTPUT_COLUMNS,
    PAYER_CONFIG,
    REMIT_NOT_FOUND_DENIAL,
    REMITTANCE_URLS,
    REQUIRED_COLUMNS,
    SELECTORS,
    STATUS_COLUMNS,
    get_selector_profile,
)
from availity_app.driver import (
    close_managed_edge_if_owned,
    ensure_edge_with_cdp,
    ensure_playwright_disconnected,
    get_live_page,
    setup_browser,
    validate_navigation_page_ready,
    _find_logged_in_page,
    _interruptible_sleep,
    is_claim_search_page,
)
from availity_app.logging_gui import log_to_gui
from availity_app.ui_control import (
    UI_STATE_IDLE,
    UI_STATE_READY,
    finish_automation_progress,
    reset_automation_progress,
    reset_ui_state,
    set_save_progress_status,
    set_save_progress_indeterminate,
    update_automation_progress,
    update_excel_row_stats,
    _prompt_user_relogin,
    _set_ui_state,
)
from availity_app.async_save import (
    flush_async_saves,
    save_sync_or_async,
    start_save_worker,
    stop_save_worker,
)
from availity_app.discovery_cache import (
    hydrate_remit_amount_paths,
    get_line_table_columns,
    persist_line_table_columns,
    persist_remit_amount_paths,
)
from availity_app.resilience import (
    BatchCircuitBreaker,
    capture_row_failure_artifact,
    classify_row_failure,
)
from availity_app.search_grouping import build_consecutive_search_groups
from availity_app.utils import (
    aggregate_invoice_visit_date_range,
    coerce_visit_date,
    empty_claim_result,
    filter_line_rows_by_visit_date,
    parse_service_date_range,
    format_invoice_number,
    normalize_date,
    normalize_date_range,
    normalize_dob,
    normalize_search_date,
    parse_amount,
    resolve_line_level_config,
    resolve_bundle_search_dates,
    resolve_service_search_dates,
    row_has_valid_dob,
    service_date_fill_plans,
    service_dates_cross_year,
    retry_with_backoff,
    normalize_invoice_column,
    normalize_patient_names,
    safe_alt_patient_id,
    safe_field,
    safe_invoice_number,
    split_date_parts,
    validate_dataframe,
    visit_date_matches_line,
)


# ============================================================================
# SECTION 7 — EXCEL / FILE OPERATIONS
# ============================================================================

def _normalize_recheck_value(value):
    """Return 'yes', 'no', or 'empty' for a Recheck cell."""
    if pd.isna(value):
        return 'empty'
    text = str(value).strip().lower()
    if text in ('', 'nan'):
        return 'empty'
    if text == 'yes':
        return 'yes'
    if text == 'no':
        return 'no'
    return text


def compute_recheck_stats(df):
    """Return (total_rows, recheck_yes_count, recheck_empty_count) from the full sheet."""
    total = len(df)
    if 'Recheck' not in df.columns:
        return total, 0, total
    kinds = df['Recheck'].map(_normalize_recheck_value)
    yes_count = int((kinds == 'yes').sum())
    empty_count = int((kinds == 'empty').sum())
    return total, yes_count, empty_count


def _is_recheck_yes(df, idx):
    """True when the row's Recheck cell is yes (case-insensitive)."""
    if 'Recheck' not in df.columns:
        return False
    return _normalize_recheck_value(df.at[idx, 'Recheck']) == 'yes'


def _claim_status_has_known_error(status_val):
    """True when Claim Status (possibly multiline) contains a known error label."""
    if not isinstance(status_val, str):
        return False
    for line in status_val.split('\n'):
        val = line.split('. ', 1)[-1].strip()
        if val in KNOWN_ERROR_STATUSES:
            return True
    return False


def _row_has_prior_error(df, idx):
    """True when the row should be retried due to a prior failure."""
    if 'LastError' in df.columns:
        last_err = df.at[idx, 'LastError']
        if not pd.isna(last_err):
            text = str(last_err).strip()
            if text and text.lower() != 'nan':
                return True

    status = str(df.at[idx, 'AutomationStatus']).strip().lower()
    if status in ('error', 'inprogress'):
        return True

    if 'Claim Status' in df.columns:
        if _claim_status_has_known_error(df.at[idx, 'Claim Status']):
            return True

    return False


def queue_error_rows_for_retry(df):
    """Reset prior-error rows to Pending so they re-enter the batch queue."""
    for idx in range(len(df)):
        if _row_has_prior_error(df, idx):
            df.at[idx, 'AutomationStatus'] = 'Pending'
    return df


def _dedupe_preserve_order(indices):
    """Return indices in first-seen order with duplicates removed."""
    seen = set()
    out = []
    for idx in indices:
        if idx not in seen:
            seen.add(idx)
            out.append(idx)
    return out


def row_should_process(df, idx):
    """True if this row should be searched/processed in the current batch.

    Prior errors always qualify (even when Recheck=no). Recheck=yes qualifies
    Done rows for manual re-run. Recheck=no skips non-error Done/Skipped rows.
    """
    if _row_has_prior_error(df, idx):
        return True
    if _is_recheck_yes(df, idx):
        return True
    if 'Recheck' in df.columns and _normalize_recheck_value(df.at[idx, 'Recheck']) == 'no':
        return False
    status = str(df.at[idx, 'AutomationStatus']).strip().lower()
    return status not in ('done', 'skipped')


def mark_recheck_skipped_rows(df):
    """Mark rows with Recheck=no as Skipped so they are not processed."""
    if 'Recheck' not in df.columns:
        return df
    for idx in range(len(df)):
        if _normalize_recheck_value(df.at[idx, 'Recheck']) == 'no':
            if not _row_has_prior_error(df, idx):
                df.at[idx, 'AutomationStatus'] = 'Skipped'
    return df


def queue_recheck_yes_rows(df):
    """Reset Done/Skipped rows with Recheck=yes so they re-enter the claim queue."""
    if 'Recheck' not in df.columns:
        return df
    for idx in range(len(df)):
        if not _is_recheck_yes(df, idx):
            continue
        status = str(df.at[idx, 'AutomationStatus']).strip().lower()
        if status in ('done', 'skipped'):
            df.at[idx, 'AutomationStatus'] = 'Pending'
    return df


def _show_excel_load_error(message):
    """Show an Excel load error dialog on the Tk main thread."""
    def _dialog():
        messagebox.showerror("Error", message)

    if state.root is not None:
        state.root.after(0, _dialog)
    else:
        log_to_gui(f"❌ {message}\n", "error")


def load_excel(file_path):
    """Load the input Excel workbook and split PatientName into Last/First.

    PatientName is expected in 'LastName, FirstName' format.
    Returns ``(dataframe, None)`` on success or ``(None, error_message)`` on failure.
    """
    try:
        df = pd.read_excel(file_path, engine='openpyxl')
        if 'PatientName' in df.columns:
            parts = df['PatientName'].str.split(',', n=1, expand=True)
            df['Last_Name']  = parts[0].str.strip()
            df['First_Name'] = parts[1].str.strip() if len(parts.columns) > 1 else ''
        df = normalize_invoice_column(df)
        return df, None
    except Exception as e:
        return None, f"Failed to load Excel file: {e}"


def load_and_prepare_excel(file_path):
    """Load Excel, validate columns, add output fields, and apply Recheck rules.

    Returns (dataframe, stats_dict) on success, or (None, None) on failure.
    stats_dict has keys: total, recheck_yes, recheck_empty.
    """
    df, load_error = load_excel(file_path)
    if df is None:
        if load_error:
            _show_excel_load_error(load_error)
        return None, None

    missing = validate_dataframe(df)
    if missing:
        log_to_gui(f"❌ Missing columns: {', '.join(missing)}\n", "error")
        log_to_gui("   Check your Excel headers and try again.\n", "error")
        return None, None

    df = add_output_columns(df)
    df = mark_recheck_skipped_rows(df)
    df = queue_recheck_yes_rows(df)
    df = queue_error_rows_for_retry(df)
    total, recheck_yes, recheck_empty = compute_recheck_stats(df)
    stats = {
        'total': total,
        'recheck_yes': recheck_yes,
        'recheck_empty': recheck_empty,
    }
    return df, stats


def add_output_columns(df):
    """Append OUTPUT_COLUMNS to df with empty defaults where they don't exist."""
    # Backward compatibility: normalize old header casing if present.
    if 'Last action taken' in df.columns and 'Last Action Taken' not in df.columns:
        df = df.rename(columns={'Last action taken': 'Last Action Taken'})
    for col in list(df.columns):
        if col != 'Recheck' and str(col).strip().lower() == 'recheck':
            df = df.rename(columns={col: 'Recheck'})
            break

    for col in OUTPUT_COLUMNS + STATUS_COLUMNS:
        if col not in df.columns:
            df[col] = ''
        else:
            # Empty Excel cells are read as float64 NaN; coerce so text results can be written.
            df[col] = df[col].fillna('').astype(object)
    df['AutomationStatus'] = df['AutomationStatus'].replace('', 'Pending').fillna('Pending')
    df['LastError'] = df['LastError'].fillna('')
    return df


def get_output_file_path(output_folder):
    """Return the single fixed output file path for all progress/final saves."""
    return os.path.join(output_folder, "Automated.xlsx")


RECHECK_LIST_SHEET = '_AvailityLists'
_AUTOFIT_MIN_COL_WIDTH = 8
_AUTOFIT_MAX_COL_WIDTH = 60
_AUTOFIT_COL_PADDING = 2
_AUTOFIT_ROW_HEIGHT_PER_LINE = 15
_AUTOFIT_MAX_ROW_HEIGHT = 150
_HEADER_FILL_COLOR = '3D5A73'
_HEADER_FONT_COLOR = 'F8FAFC'
_HEADER_ROW_HEIGHT = 22
_LARGE_FILE_ROW_THRESHOLD = 500
_LARGE_FILE_SAVE_INTERVAL = 5
_LARGE_AUTOFIT_ROW_THRESHOLD = 500
_AUTOFIT_SAMPLE_ROWS = 100
_OUTPUT_FORMATTING_ENABLED = False


def _cell_display_length(value):
    """Return the longest line length for a cell value (handles multiline text)."""
    if value is None:
        return 0
    text = str(value)
    if text.lower() == 'nan':
        return 0
    lines = text.split('\n') if text else ['']
    return max(len(line) for line in lines)


def _autofit_worksheet(ws, *, max_data_row=None):
    """Set wrap alignment, column widths, and row heights so content stays visible."""
    from openpyxl.styles import Alignment
    from openpyxl.utils import get_column_letter

    last_row = ws.max_row
    if max_data_row is not None:
        last_row = min(last_row, max_data_row)

    wrap = Alignment(wrap_text=True, vertical='top')
    for row in ws.iter_rows(min_row=2, max_row=last_row):
        for cell in row:
            cell.alignment = wrap

    col_max = {}
    for row in ws.iter_rows(max_row=last_row):
        for cell in row:
            col_idx = cell.column
            length = _cell_display_length(cell.value)
            col_max[col_idx] = max(col_max.get(col_idx, 0), length)

    for col_idx, max_len in col_max.items():
        width = min(
            max(max_len + _AUTOFIT_COL_PADDING, _AUTOFIT_MIN_COL_WIDTH),
            _AUTOFIT_MAX_COL_WIDTH,
        )
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    for row_idx in range(2, last_row + 1):
        max_lines = 1
        for cell in ws[row_idx]:
            if cell.value is None:
                continue
            text = str(cell.value)
            if text.lower() == 'nan':
                continue
            max_lines = max(max_lines, len(text.split('\n')))
        if max_lines > 1:
            ws.row_dimensions[row_idx].height = min(
                _AUTOFIT_ROW_HEIGHT_PER_LINE * max_lines,
                _AUTOFIT_MAX_ROW_HEIGHT,
            )


def _autofit_worksheet_lite(ws, *, sample_rows=_AUTOFIT_SAMPLE_ROWS):
    """Column widths from header + first N data rows only (large-file fast path)."""
    from openpyxl.utils import get_column_letter

    last_row = min(ws.max_row, 1 + sample_rows)
    col_max = {}
    for row in ws.iter_rows(min_row=1, max_row=last_row):
        for cell in row:
            col_idx = cell.column
            length = _cell_display_length(cell.value)
            col_max[col_idx] = max(col_max.get(col_idx, 0), length)

    for col_idx, max_len in col_max.items():
        width = min(
            max(max_len + _AUTOFIT_COL_PADDING, _AUTOFIT_MIN_COL_WIDTH),
            _AUTOFIT_MAX_COL_WIDTH,
        )
        ws.column_dimensions[get_column_letter(col_idx)].width = width


def _apply_multiline_row_formatting(ws, *, row_indices=None):
    """Enable wrap text and row height on rows with newline-separated cell values.

    When ``row_indices`` is provided, only those 1-based Excel row numbers are
    scanned (fast path for large files with few multi-claim rows).
    """
    from openpyxl.styles import Alignment

    wrap = Alignment(wrap_text=True, vertical='top')
    if row_indices is None:
        row_indices = range(2, ws.max_row + 1)
    for row_idx in row_indices:
        if row_idx < 2 or row_idx > ws.max_row:
            continue
        max_lines = 1
        row_has_multiline = False
        for cell in ws[row_idx]:
            if cell.value is None:
                continue
            text = str(cell.value)
            if text.lower() == 'nan':
                continue
            line_count = len(text.split('\n'))
            if line_count > 1:
                row_has_multiline = True
                cell.alignment = wrap
                max_lines = max(max_lines, line_count)
        if row_has_multiline:
            ws.row_dimensions[row_idx].height = min(
                _AUTOFIT_ROW_HEIGHT_PER_LINE * max_lines,
                _AUTOFIT_MAX_ROW_HEIGHT,
            )


def _dataframe_multiline_excel_rows(df):
    """Return 1-based Excel row numbers (incl. header offset) with newline cells."""
    rows = []
    for idx in range(len(df)):
        row = df.iloc[idx]
        for col in df.columns:
            val = row[col]
            if isinstance(val, str) and '\n' in val:
                rows.append(idx + 2)
                break
    return rows


def _style_header_row(ws):
    """Apply minimalist fill, font, and alignment to the header row."""
    from openpyxl.styles import Alignment, Font, PatternFill

    header_fill = PatternFill(
        start_color=_HEADER_FILL_COLOR,
        end_color=_HEADER_FILL_COLOR,
        fill_type='solid',
    )
    header_font = Font(
        name='Calibri',
        size=11,
        bold=True,
        color=_HEADER_FONT_COLOR,
    )
    header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    ws.row_dimensions[1].height = _HEADER_ROW_HEIGHT
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_align
    ws.freeze_panes = 'A2'


def _clear_recheck_cells(ws, recheck_col, last_row):
    """Blank Recheck data cells while leaving validation for the dropdown."""
    if recheck_col is None or last_row < 2:
        return
    for row_idx in range(2, last_row + 1):
        ws.cell(row=row_idx, column=recheck_col).value = None


def _find_recheck_column_index(worksheet):
    """Return 1-based column index for the Recheck header, or None."""
    for cell in worksheet[1]:
        value = cell.value
        if value is not None and str(value).strip().lower() == 'recheck':
            return cell.column
    return None


def _recheck_list_formula(workbook):
    """Hidden sheet with yes/no; range reference works reliably in Excel."""
    from openpyxl.utils import quote_sheetname

    if RECHECK_LIST_SHEET in workbook.sheetnames:
        del workbook[RECHECK_LIST_SHEET]
    list_ws = workbook.create_sheet(RECHECK_LIST_SHEET)
    list_ws['A1'] = 'yes'
    list_ws['A2'] = 'no'
    list_ws.sheet_state = 'hidden'
    sheet_ref = quote_sheetname(RECHECK_LIST_SHEET)
    return f'{sheet_ref}!$A$1:$A$2'


def _finalize_output_workbook(
    workbook_path,
    *,
    data_row_count=None,
    recheck_col=None,
    clear_recheck_values=False,
    multiline_row_indices=None,
):
    """Post-process Automated.xlsx: optional Recheck clear, autofit, dropdown."""
    from openpyxl import load_workbook
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.datavalidation import DataValidation

    set_save_progress_status("Formatting workbook…")
    try:
        wb = load_workbook(workbook_path)
        ws = wb.worksheets[0]
        if recheck_col is None:
            recheck_col = _find_recheck_column_index(ws)

        last_row = ws.max_row
        if data_row_count:
            last_row = max(last_row, data_row_count + 1)

        if clear_recheck_values:
            _clear_recheck_cells(ws, recheck_col, last_row)

        large_output = bool(data_row_count and data_row_count > _LARGE_AUTOFIT_ROW_THRESHOLD)
        if large_output:
            if not state._large_autofit_notice_logged:
                log_to_gui(
                    "  ℹ️ Large output — skipped full autofit "
                    "(multiline rows still formatted).\n",
                    "info",
                )
                state._large_autofit_notice_logged = True
            _autofit_worksheet_lite(ws)
        else:
            _autofit_worksheet(ws)
        _apply_multiline_row_formatting(
            ws,
            row_indices=multiline_row_indices,
        )
        _style_header_row(ws)

        if recheck_col is not None and last_row >= 2:
            col_letter = get_column_letter(recheck_col)
            cell_range = f'{col_letter}2:{col_letter}{last_row}'

            ws.data_validations.dataValidation.clear()

            list_formula = _recheck_list_formula(wb)
            dv = DataValidation(
                type='list',
                formula1=list_formula,
                allow_blank=True,
                # openpyxl: False => Excel shows the dropdown arrow (inverted attr).
                showDropDown=False,
                showInputMessage=True,
                showErrorMessage=True,
            )
            dv.error = 'Select yes or no, or leave blank for new rows.'
            dv.errorTitle = 'Invalid Recheck'
            dv.prompt = 'yes = reprocess row; no = skip; blank = new row'
            dv.promptTitle = 'Recheck'
            dv.add(cell_range)
            ws.add_data_validation(dv)

        wb.active = ws
        wb.save(workbook_path)
    except Exception as e:
        log_to_gui(f'  ⚠️ Output workbook formatting failed: {e}\n', 'error')
        raise


def _acquire_save_lock():
    """Block until the save lock is free; show status if another save is in progress."""
    if state.save_lock.locked():
        set_save_progress_status("Waiting for save to finish…")
    state.save_lock.acquire()


def _should_save_progress(row_count, rows_since_last_save):
    """Return True when a mid-batch progress save should run."""
    if row_count <= _LARGE_FILE_ROW_THRESHOLD:
        return True
    return rows_since_last_save >= _LARGE_FILE_SAVE_INTERVAL


def save_progress_results(df, output_folder, *, force=False):
    """Throttled mid-batch save; always saves when ``force`` is True."""
    if force or _should_save_progress(len(df), state.rows_since_last_save):
        save_sync_or_async(df, output_folder, finalize=False, force_async=not force)
        state.rows_since_last_save = 0
    else:
        state.rows_since_last_save += 1


def _write_results_file(df, output_folder, *, finalize=True):
    """Write df to Automated.xlsx (used by sync and async save paths)."""
    _acquire_save_lock()
    if finalize:
        status = (
            "Saving and formatting results…"
            if _OUTPUT_FORMATTING_ENABLED
            else "Saving results…"
        )
        set_save_progress_status(status)
    else:
        set_save_progress_status("Saving progress…")
    path = get_output_file_path(output_folder)
    fd, temp_path = tempfile.mkstemp(suffix='.xlsx')
    os.close(fd)
    try:
        export_df = df.drop(columns=['Last_Name', 'First_Name'], errors='ignore')
        output_cols = [c for c in OUTPUT_COLUMNS if c in export_df.columns]
        status_cols = [c for c in STATUS_COLUMNS if c in export_df.columns]
        base_cols = [c for c in export_df.columns if c not in output_cols + status_cols]
        export_df = export_df[base_cols + output_cols + status_cols]
        if 'Recheck' in export_df.columns:
            export_df = export_df.copy()
            export_df['Recheck'] = ''
        export_df.to_excel(temp_path, index=False, engine='openpyxl')
        if finalize and _OUTPUT_FORMATTING_ENABLED:
            recheck_col = None
            if 'Recheck' in export_df.columns:
                recheck_col = int(export_df.columns.get_loc('Recheck')) + 1
            multiline_rows = _dataframe_multiline_excel_rows(export_df)
            _finalize_output_workbook(
                temp_path,
                data_row_count=len(export_df),
                recheck_col=recheck_col,
                clear_recheck_values=True,
                multiline_row_indices=multiline_rows or None,
            )
        os.replace(temp_path, path)
        temp_path = None
        log_to_gui(f"  💾 Saved: {path}\n", "info")
        return path
    except OSError as e:
        log_to_gui(
            f"  ❌ Save error: {e}\n"
            "     Close Automated.xlsx if it is open in Excel, then try again.\n",
            "error",
        )
        return None
    except Exception as e:
        log_to_gui(f"  ❌ Save error: {e}\n", "error")
        return None
    finally:
        if temp_path and os.path.exists(temp_path):
            try:
                os.remove(temp_path)
            except OSError:
                pass
        state.save_lock.release()


def save_results(df, output_folder, *, finalize=True):
    """Write df to one fixed Excel workbook in output_folder.

    Last_Name and First_Name are excluded from the output because they are
    derived from PatientName and are not needed by downstream consumers.
    Serialized via ``save_lock`` so the GUI-thread Stop save and the worker
    row save can't interleave the same write call.

    When ``finalize`` is True and ``_OUTPUT_FORMATTING_ENABLED`` is True,
    autofits columns/rows, styles the header row, clears Recheck values in
    the file only, and applies the yes/no Recheck dropdown. Mid-batch progress
    saves use ``finalize=False`` for speed.

    Returns the saved file path, or None on failure.
    """
    flush_async_saves(timeout=120.0)
    return _write_results_file(df, output_folder, finalize=finalize)


def _run_finalize_save(df, output_folder, *, timeout=300.0):
    """Run formatted save on a background thread; block caller until done."""
    done = threading.Event()
    result = {'path': None}

    def _worker():
        try:
            set_save_progress_indeterminate(True)
            set_save_progress_status(
                "Saving and formatting results…"
                if _OUTPUT_FORMATTING_ENABLED
                else "Saving results…"
            )
            result['path'] = save_results(df, output_folder, finalize=True)
        finally:
            set_save_progress_indeterminate(False)
            done.set()

    thread = threading.Thread(
        target=_worker, name='availity-final-save', daemon=True,
    )
    thread.start()
    if not done.wait(timeout=timeout):
        log_to_gui(
            "  ⚠️ Final save did not finish in time — check Automated.xlsx.\n",
            "error",
        )
    thread.join(timeout=1.0)
    return result['path']


def write_row_result(df, row_index, result_dict):
    """Apply every key/value in result_dict to the given dataframe row in-place."""
    for field, value in result_dict.items():
        if field in df.columns and pd.api.types.is_numeric_dtype(df[field].dtype):
            df[field] = df[field].astype(object)
        df.at[row_index, field] = value


def safe_mark_row_status(df, row_index, status, err=''):
    """Safely update row processing status columns."""
    for col in STATUS_COLUMNS:
        if col in df.columns and pd.api.types.is_numeric_dtype(df[col].dtype):
            df[col] = df[col].astype(object)
    df.at[row_index, 'AutomationStatus'] = status
    df.at[row_index, 'LastError'] = err


# ============================================================================
# SECTION 8 — PAGE NAVIGATION
# ============================================================================

_PLAIN_INPUT_DATE_MODES = frozenset({'legacy_input', 'plain_input'})

_PAGE_READY_TIMEOUT_MS = 30000


def _uses_plain_date_inputs(sel):
    """True when search dates are filled via plain text inputs (not MUI pickers)."""
    return sel.get('date_fill_mode') in _PLAIN_INPUT_DATE_MODES


def _resolve_page(page):
    """Return ``page`` when still open, else fall back to any live browser tab."""
    try:
        if page is not None and not page.is_closed():
            return page
    except Exception:
        pass
    return get_live_page(page)


def _claim_search_iframe(page, payer_config=None):
    """Return a FrameLocator for the claim-search iframe after attach check."""
    page = _resolve_page(page)
    sel = get_selector_profile(payer_config)
    page.wait_for_selector(
        sel['iframe'], state='attached', timeout=_PAGE_READY_TIMEOUT_MS,
    )
    return page.frame_locator(sel['iframe'])


def wait_for_claim_results_ready(page, payer_config=None, quiet=False,
                                 timeout_ms=None):
    """Block until the claim-search iframe and #claimsTable are visible.

    Use after the Results breadcrumb — not the search form.
    """
    timeout_ms = timeout_ms or _PAGE_READY_TIMEOUT_MS
    page = _resolve_page(page)
    if not is_claim_search_page(page):
        raise RuntimeError(
            "Browser tab is not on claim search — navigation shell only "
            f"({page.url})"
        )
    if not quiet:
        log_to_gui("  → Waiting for results table...\n")
    iframe = _claim_search_iframe(page, payer_config)
    iframe.locator(SELECTORS['results_table']).wait_for(
        state='visible', timeout=timeout_ms,
    )
    if not quiet:
        log_to_gui("  ✓ Results table ready\n", "success")
    return iframe


def _is_on_claim_details_page(page):
    """True when the outer URL is on a claim detail route."""
    url = (page.url or '').lower()
    return 'standard-details' in url or 'claim-details' in url


def _results_table_is_visible(page, payer_config=None, timeout_ms=5000):
    """True when #claimsTable is visible inside the search iframe."""
    try:
        iframe = _claim_search_iframe(page, payer_config)
        iframe.locator(SELECTORS['results_table']).first.wait_for(
            state='visible', timeout=timeout_ms,
        )
        return True
    except Exception:
        return False


def _rerun_claim_search(page, row_data, payer_config):
    """Return to the search form and re-submit to rebuild #claimsTable."""
    log_to_gui("  → Re-running claim search to restore results...\n", "info")
    # MUI service-date range pickers reject in-place edits after a prior search;
    # always unload/reload so the next fill starts on a clean form.
    page = _reload_and_navigate(page, payer_config)
    iframe = wait_for_page_ready(page, payer_config=payer_config)
    if not fill_search_form(page, iframe, row_data, payer_config):
        raise RuntimeError('Re-search form fill failed')
    outcome = submit_search_and_wait(iframe)
    if outcome != 'results':
        raise RuntimeError(f'Re-search did not return results: {outcome}')
    return page, iframe


def _return_to_results_view(page, payer_config=None):
    """Return to the results list via browser back or Results breadcrumb."""
    page = _resolve_page(page)
    results_timeout_ms = 12000

    if _is_on_claim_details_page(page):
        try:
            log_to_gui("  → Browser back to results list...\n")
            page.go_back(wait_until='domcontentloaded', timeout=20000)
            if _results_table_is_visible(page, payer_config, timeout_ms=results_timeout_ms):
                return wait_for_claim_results_ready(
                    page, payer_config=payer_config, quiet=True,
                    timeout_ms=results_timeout_ms,
                )
        except Exception:
            pass

    sel = get_selector_profile(payer_config)
    iframe = page.frame_locator(sel['iframe'])
    results_link = iframe.locator(SELECTORS['breadcrumb_results'])
    if results_link.first.is_visible(timeout=5000):
        results_link.first.click(timeout=15000)
        page.wait_for_load_state('domcontentloaded', timeout=45000)
        if _results_table_is_visible(page, payer_config, timeout_ms=results_timeout_ms):
            return wait_for_claim_results_ready(
                page, payer_config=payer_config, quiet=True,
                timeout_ms=results_timeout_ms,
            )

    raise RuntimeError(
        f'Results table did not appear after back navigation ({page.url})'
    )


def _wait_for_search_form_in_iframe(iframe, sel, timeout_ms=15000):
    """Wait until any known search-form control from ``sel`` is visible."""
    probes = [sel['member_id']]
    if _uses_plain_date_inputs(sel):
        probes.extend(k for k in (sel.get('dob'), sel.get('date_from')) if k)
    probes.append(sel['submit_btn'])

    combined = iframe.locator(probes[0])
    for selector in probes[1:]:
        combined = combined.or_(iframe.locator(selector))

    last_error = None
    try:
        combined.first.wait_for(state='visible', timeout=timeout_ms)
        for selector in probes:
            try:
                if iframe.locator(selector).first.is_visible():
                    return selector
            except Exception:
                continue
        return probes[0]
    except Exception as e:
        last_error = e

    deadline = time.time() + (timeout_ms / 1000.0)
    for selector in probes:
        remaining_ms = max(500, int((deadline - time.time()) * 1000))
        if remaining_ms <= 0:
            break
        try:
            iframe.locator(selector).wait_for(state='visible', timeout=remaining_ms)
            return selector
        except Exception as e:
            last_error = e
    raise last_error or Exception("Claim search form did not become ready")


def wait_for_page_ready(page, payer_config=None, quiet=False):
    """Block until the claim-search iframe and form fields are visible.

    Uses the payer's ``selector_profile`` (legacy vs MUI) when probing for
    readiness.  Raises on failure so retry_with_backoff() can retry the step.
    Returns the FrameLocator for the search iframe.

    ``quiet``: when True, skip routine log lines (used during session probes).
    """
    page = _resolve_page(page)
    sel = get_selector_profile(payer_config)
    if not quiet:
        log_to_gui("  → Waiting for iframe...\n")
    if not is_claim_search_page(page):
        raise RuntimeError(
            "Browser tab is not on claim search — navigation shell only "
            f"({page.url})"
        )
    try:
        page.wait_for_selector(
            sel['iframe'], state='attached', timeout=_PAGE_READY_TIMEOUT_MS,
        )
    except Exception as e:
        raise RuntimeError(
            f"Claim search iframe not attached ({page.url}): {e}"
        ) from e
    time.sleep(1)
    iframe = page.frame_locator(sel['iframe'])
    try:
        matched = _wait_for_search_form_in_iframe(
            iframe, sel, timeout_ms=_PAGE_READY_TIMEOUT_MS,
        )
    except Exception as e:
        raise RuntimeError(
            f"Claim search form did not become ready ({page.url}): {e}"
        ) from e
    if not quiet:
        log_to_gui(f"  ✓ Page ready ({matched})\n", "success")
    return iframe


def navigate_to_payer_page(page, payer_config):
    """Navigate to the payer's claim-search URL and click the HIPAA tab if required.

    Separating navigation from page-readiness lets each step use its own
    retry budget independently.

    Args:
        page:         Playwright Page object.
        payer_config: Entry from PAYER_CONFIG for the selected payer.
    """
    page = retry_with_backoff(
        "Acquire live page",
        lambda: get_live_page(page),
        attempts=3,
    )
    retry_with_backoff(
        "Navigate to search page",
        lambda: page.goto(payer_config['url'], wait_until='domcontentloaded', timeout=45000),
        attempts=3,
    )
    # The SPA's client-side routing settles implicitly: every branch below (and
    # the caller's wait_for_page_ready) blocks on a visible form control.
    sel = get_selector_profile(payer_config)
    if payer_config['uses_hipaa_tab']:
        def _click_hipaa():
            live_page = get_live_page(page)
            hipaa_tab = live_page.frame_locator(sel['iframe']).locator(sel['hipaa_tab'])
            hipaa_tab.first.wait_for(state='visible', timeout=20000)
            hipaa_tab.first.click(timeout=15000)
            return live_page

        page = retry_with_backoff("Select HIPAA tab", _click_hipaa, attempts=3)
        log_to_gui("  ✓ HIPAA tab selected\n", "success")
        iframe = get_live_page(page).frame_locator(sel['iframe'])
        _wait_for_search_form_in_iframe(iframe, sel, timeout_ms=20000)
    elif _uses_plain_date_inputs(sel):
        # Plain-input payers without a separate HIPAA tab still need the form.
        iframe = get_live_page(page).frame_locator(sel['iframe'])
        _wait_for_search_form_in_iframe(iframe, sel, timeout_ms=20000)
    else:
        iframe = get_live_page(page).frame_locator(sel['iframe'])
        _wait_for_search_form_in_iframe(iframe, sel, timeout_ms=20000)
    return page


def _verify_payer_search_session(page, payer_config, *, quiet=False):
    """Confirm the Availity session can open the payer claim search (not a stale nav tab).

    Returns ``(True, page)`` when the search iframe and member-id field are reachable,
    else ``(False, page)`` (``page`` is best-effort live page for continued waiting).
    """
    if payer_config is None:
        return True, page
    try:
        wp = navigate_to_payer_page(page, payer_config)
        wait_for_page_ready(wp, payer_config=payer_config, quiet=quiet)
        return True, wp
    except Exception as e:
        log_to_gui(f"  ⚠️ Session check failed: {e}\n", "error")
        try:
            return False, get_live_page(page)
        except Exception:
            return False, page


# ============================================================================
# SECTION 9 — FORM INTERACTION
# ============================================================================

def _ensure_checkbox_checked(iframe, selector, label='checkbox'):
    """Check a form checkbox when present and not already selected."""
    box = iframe.locator(selector).first
    box.wait_for(state='visible', timeout=10000)
    if not box.is_checked():
        box.check(force=True)
        log_to_gui(f"  ✓ Checked {label}\n", "success")


def _fill_plain_date_inputs(iframe, sel, dob, start_date, end_date):
    """Fill plain text date inputs (legacy HIPAA-tab or Aetna DateInput fields)."""
    if dob:
        dob_input = iframe.locator(sel['dob'])
        dob_input.click()
        dob_input.fill(dob)
    if start_date:
        from_input = iframe.locator(sel['date_from'])
        from_input.click()
        from_input.fill(start_date)
    if end_date:
        to_input = iframe.locator(sel['date_to'])
        to_input.click()
        to_input.fill(end_date)


_form_fill_last_error = ''


def _is_service_date_form_error(error_msg):
    """True when a form-fill failure is likely caused by service-date entry."""
    markers = (
        'Service To read-back',
        'Service From read-back',
        'Failed to fill Service',
        'Service date',
        'date group with',
    )
    return any(marker in (error_msg or '') for marker in markers)


def fill_search_form(page, iframe, row_data, payer_config=None, *, service_start=None, service_end=None):
    """Populate the claim search form with data from a CSV row.

    Uses safe_field() throughout to guard against NaN values.
    Date fields use the payer's selector profile (MUI spinbuttons or legacy inputs).

    Optional ``service_start`` / ``service_end`` override row StartDate/EndDate
    (used for Fidelis bundle search retry on a reloaded form).

    Returns True on success, False if a required field is empty or a
    Playwright action raises.
    """
    global _form_fill_last_error
    _form_fill_last_error = ''
    try:
        log_to_gui("  → Filling search form...\n")
        sel = get_selector_profile(payer_config)
        member_id = safe_alt_patient_id(row_data, 'AltPatientID')
        raw_last = safe_field(row_data, 'Last_Name')
        raw_first = safe_field(row_data, 'First_Name')
        last_name, first_name = normalize_patient_names(raw_last, raw_first)
        if last_name != raw_last or first_name != raw_first:
            log_to_gui(
                f"  ℹ️ Name normalized: {raw_last!r}, {raw_first!r} → "
                f"{last_name!r}, {first_name!r}\n",
                "info",
            )
        dob = normalize_dob(safe_field(row_data, 'DOB'))
        if service_start and service_end:
            start_date, end_date = service_start, service_end
            log_to_gui(
                f"  ℹ️ Service search dates (bundle retry): {start_date} → {end_date}\n",
                "info",
            )
        else:
            start_date, end_date = resolve_service_search_dates(row_data, safe_field)
        if not start_date and not end_date:
            visit_raw = safe_field(row_data, 'VisitDate')
            start_raw = safe_field(row_data, 'StartDate')
            end_raw = safe_field(row_data, 'EndDate')
            raise ValueError(
                "No valid service dates: StartDate, EndDate, and VisitDate are "
                f"missing or invalid (StartDate={start_raw!r}, EndDate={end_raw!r}, "
                f"VisitDate={visit_raw!r})"
            )
        if not normalize_search_date(safe_field(row_data, 'StartDate')) or not normalize_search_date(
            safe_field(row_data, 'EndDate')
        ):
            if not (service_start and service_end):
                log_to_gui(
                    f"  ℹ️ Service search dates: {start_date} → {end_date} "
                    f"(invalid StartDate/EndDate — using VisitDate when available)\n",
                    "info",
                )

        # AltPatientID and Last_Name are required for a meaningful search
        if not member_id:
            raise ValueError("AltPatientID is empty or missing")
        if not last_name:
            raise ValueError("Last_Name is empty or missing")
        if not _uses_plain_date_inputs(sel) and not dob:
            raise ValueError("DOB is empty or missing")

        if sel.get('patient_is_subscriber'):
            _ensure_checkbox_checked(
                iframe, sel['patient_is_subscriber'], 'Patient is subscriber',
            )

        iframe.locator(sel['member_id']).fill(member_id)
        iframe.locator(sel['last_name']).fill(last_name)
        iframe.locator(sel['first_name']).fill(first_name)
        if state.current_payer_name != 'Anthem BCBS':
            iframe.locator(sel['provider_npi']).fill(NPI)
        if _uses_plain_date_inputs(sel):
            _fill_plain_date_inputs(iframe, sel, dob, start_date, end_date)
        else:
            if dob:
                fill_mui_date_picker(page, iframe, sel['patient_dob_label'], dob)
            if start_date and end_date:
                fill_date_range(
                    page, iframe, start_date, end_date,
                    from_label=sel['service_from_date_label'],
                    to_label=sel['service_to_date_label'],
                )
            elif start_date:
                fill_mui_date_picker(
                    page, iframe, sel['service_from_date_label'], start_date
                )
            elif end_date:
                fill_mui_date_picker(
                    page, iframe, sel['service_to_date_label'], end_date
                )
        return True
    except Exception as e:
        _form_fill_last_error = str(e)
        log_to_gui(f"  ❌ Form fill error: {e}\n", "error")
        return False


def _handle_form_fill_failure(
    page, iframe, df, row_index, row_data, output_folder, payer_config,
    *, service_start=None, service_end=None,
):
    """Handle form fill failure; reload once on service-date errors.

    Returns ``(outcome, page, iframe)`` where outcome is:
      ``continue`` — retry succeeded, proceed to search
      ``row_done`` — row handled (e.g. DOB skip)
      ``failed``   — row marked error
    """
    if not row_has_valid_dob(row_data, safe_field):
        page = _finish_row_without_claims(
            page, df, row_index, output_folder, payer_config,
            claim_status=DOB_NOT_FOUND_STATUS,
            automation_status='Skipped',
            success_log=f"  ✓ Row {row_index + 1} skipped ({DOB_NOT_FOUND_STATUS})",
        )
        return 'row_done', page, iframe

    error = _form_fill_last_error
    if _is_service_date_form_error(error):
        start_date, end_date = resolve_service_search_dates(row_data, safe_field)
        if service_start and service_end:
            start_date, end_date = service_start, service_end
        log_to_gui(
            "  → Form fill date error — reloading search form and retrying once...\n",
            "info",
        )
        try:
            page = _reload_and_navigate(page, payer_config)
            iframe = retry_with_backoff(
                "Wait for page ready (form date retry)",
                lambda: wait_for_page_ready(page, payer_config=payer_config),
                attempts=3,
            )
            if fill_search_form(
                page, iframe, row_data, payer_config,
                service_start=service_start, service_end=service_end,
            ):
                return 'continue', page, iframe
            error = _form_fill_last_error or error
        except Exception as retry_error:
            error = str(retry_error)
            log_to_gui(f"  ❌ Form date retry failed: {retry_error}\n", "error")

        capture_row_failure_artifact(
            page, output_folder, row_index,
            {
                'error': error,
                'stage': 'form_fill_date',
                'payer': state.current_payer_name or '',
                'invoice': safe_invoice_number(row_data, 'InvoiceNumber', ''),
                'start_date': start_date,
                'end_date': end_date,
                'cross_year': service_dates_cross_year(start_date, end_date),
            },
        )

    write_row_result(df, row_index, empty_claim_result('Form error'))
    safe_mark_row_status(df, row_index, 'Error', error or 'Form error')
    save_progress_results(df, output_folder, force=True)
    return 'failed', page, iframe


def _row_used_start_end_dates(row_data):
    """True when both StartDate and EndDate are valid for the search form."""
    return bool(
        normalize_search_date(safe_field(row_data, 'StartDate'))
        and normalize_search_date(safe_field(row_data, 'EndDate'))
    )


def _try_bundle_search_date_fallback(
    page, iframe, df, invoice_number, row_data, payer_config,
    initial_start, initial_end, *, already_attempted=False,
):
    """Retry search with min–max VisitDate bundle bounds when configured.

    Reloads a clean search form (MUI range pickers reject in-place date changes
    after a prior search). Returns ``(outcome, page, iframe)`` or ``None``.
    """
    if already_attempted:
        return None
    if not payer_config or not payer_config.get('bundle_search_date_fallback'):
        return None
    if not _row_used_start_end_dates(row_data):
        return None

    bundle_start, bundle_end = resolve_bundle_search_dates(df, invoice_number)
    if not bundle_start or not bundle_end:
        return None
    if (bundle_start, bundle_end) == (initial_start, initial_end):
        return None

    visit_rows = sum(
        1
        for _, row in df.iterrows()
        if format_invoice_number(row.get('InvoiceNumber', ''), '')
        == format_invoice_number(invoice_number)
    )
    log_to_gui(
        f"  → Bundle date search retry: {bundle_start} → {bundle_end} "
        f"({visit_rows} collection row(s))\n",
        "info",
    )
    try:
        page = _reload_and_navigate(page, payer_config)
        iframe = retry_with_backoff(
            "Wait for page ready (bundle retry)",
            lambda: wait_for_page_ready(page, payer_config=payer_config),
            attempts=3,
        )
        if not fill_search_form(
            page, iframe, row_data, payer_config,
            service_start=bundle_start, service_end=bundle_end,
        ):
            log_to_gui("  ⚠️ Bundle retry form fill failed\n", "error")
            return 'failed', page, iframe
        outcome = submit_search_and_wait(iframe)
        return outcome, page, iframe
    except Exception as e:
        log_to_gui(f"  ⚠️ Bundle date search retry failed: {e}\n", "error")
        return 'failed', page, iframe


# Populated during submit_search_and_wait so row completion never re-queries a busy iframe.
_last_search_alert_message = ''

# Availity info banners that are always on the form — not a search result.
_INFORMATIONAL_ALERT_PHRASES = (
    'supports only commercial',
    'please use the hipaa',
    'lines of business',
)


def _is_informational_payer_alert(message):
    """True for static guidance banners that must not end the search wait."""
    if not message:
        return False
    lower = message.lower()
    phrases = _INFORMATIONAL_ALERT_PHRASES
    payer_config = state.selected_payer_config
    if payer_config:
        extra = payer_config.get('informational_alert_phrases') or ()
        phrases = phrases + tuple(extra)
    return any(phrase in lower for phrase in phrases)


def _read_visible_payer_alert_text(iframe, *, skip_informational=False):
    """Return the first visible MUI alert message inside the search iframe."""
    alert = iframe.locator(SELECTORS['payer_search_alert'])
    try:
        count = alert.count()
    except Exception:
        return ''
    for i in range(count):
        item = alert.nth(i)
        try:
            if not item.is_visible():
                continue
            msg = item.locator('.MuiAlert-message').first.text_content(timeout=500)
            if not msg:
                continue
            msg = msg.strip()
            if skip_informational and _is_informational_payer_alert(msg):
                continue
            return msg
        except Exception:
            continue
    return ''


def _classify_payer_alert_text(message):
    """Map payer alert copy to a search outcome bucket."""
    if not message:
        return None
    lower = message.lower()
    if 'subscriber' in lower and 'not found' in lower:
        return 'invalid_member_id'
    if 'member' in lower and ('not found' in lower or 'invalid' in lower):
        return 'invalid_member_id'
    if 'acknowledgement/not found' in lower:
        return 'not_found'
    if 'could not find any results' in lower:
        return 'not_found'
    if 'not found' in lower:
        return 'not_found'
    if 'taking longer than anticipated' in lower or 'please wait while we retrieve' in lower:
        return 'transient_wait'
    return 'payer_error'


def _log_payer_alert(message):
    if message:
        log_to_gui(f"  ℹ️ Payer alert: {message}\n", "info")


def is_invalid_member_id_alert(iframe):
    """True when Availity shows subscriber/member id not found instead of results."""
    global _last_search_alert_message
    if iframe.locator(SELECTORS['invalid_member_id_alert']).count() > 0:
        alert = iframe.locator(SELECTORS['invalid_member_id_alert']).first
        if alert.is_visible():
            message = _read_visible_payer_alert_text(iframe, skip_informational=True)
            if message:
                _last_search_alert_message = message
                _log_payer_alert(message)
            return True
    message = _read_visible_payer_alert_text(iframe, skip_informational=True)
    if _classify_payer_alert_text(message) == 'invalid_member_id':
        _last_search_alert_message = message
        _log_payer_alert(message)
        return True
    return False


def is_claim_not_found_alert(iframe):
    """True when Availity shows payer acknowledgement that claim is not in adjudication."""
    ack = iframe.locator(SELECTORS['claim_not_found_ack'])
    if ack.count() > 0 and ack.first.is_visible():
        _log_payer_alert(_read_visible_payer_alert_text(iframe, skip_informational=True))
        return True
    warn = iframe.locator(SELECTORS['claim_not_found_warning'])
    if warn.count() > 0 and warn.first.is_visible():
        _log_payer_alert(_read_visible_payer_alert_text(iframe, skip_informational=True))
        return True
    message = _read_visible_payer_alert_text(iframe, skip_informational=True)
    if _classify_payer_alert_text(message) == 'not_found':
        _log_payer_alert(message)
        return True
    return False


def _resolve_payer_error_status(iframe=None):
    """Claim Status text for a generic payer alert after search."""
    message = _last_search_alert_message
    if not message and iframe is not None:
        message = _read_visible_payer_alert_text(iframe, skip_informational=True)
    if not message:
        return PAYER_SEARCH_ERROR_STATUS
    if len(message) > 120:
        return f'{PAYER_SEARCH_ERROR_STATUS}: {message[:117]}...'
    return f'{PAYER_SEARCH_ERROR_STATUS}: {message}'


def _resolve_search_outcome(iframe):
    """Return search outcome: invalid_member_id, not_found, payer_error, results, or failed."""
    global _last_search_alert_message
    if is_invalid_member_id_alert(iframe):
        return 'invalid_member_id'
    if is_claim_not_found_alert(iframe):
        return 'not_found'
    message = _read_visible_payer_alert_text(iframe, skip_informational=True)
    _last_search_alert_message = message
    if message:
        bucket = _classify_payer_alert_text(message)
        if bucket in ('payer_error', 'transient_wait'):
            _log_payer_alert(message)
            return 'payer_error'
    if iframe.locator(SELECTORS['results_table']).is_visible():
        return 'results'
    return 'failed'


def _search_response_ready(iframe):
    """True when the search has produced results or a blocking payer alert."""
    try:
        if iframe.locator(SELECTORS['results_table']).is_visible():
            return True
    except Exception:
        pass
    for selector in (
        SELECTORS['invalid_member_id_alert'],
        SELECTORS['claim_not_found_ack'],
        SELECTORS['claim_not_found_warning'],
    ):
        try:
            loc = iframe.locator(selector)
            if loc.count() > 0 and loc.first.is_visible():
                return True
        except Exception:
            continue
    message = _read_visible_payer_alert_text(iframe, skip_informational=True)
    if message and _classify_payer_alert_text(message) in (
        'invalid_member_id', 'not_found', 'payer_error',
    ):
        return True
    return False


def _wait_for_search_response(iframe, timeout_ms=15000, max_timeout_ms=45000):
    """Poll until results or a terminal payer alert appear.

    A transient "taking longer than anticipated" alert extends the wait in
    3s steps up to ``max_timeout_ms`` (default 45s total from search start).
    """
    start = time.time()
    deadline = start + (timeout_ms / 1000.0)
    max_deadline = start + (max_timeout_ms / 1000.0)
    logged_transient = False

    while time.time() < deadline:
        if not state.is_running:
            raise TimeoutError('Search wait interrupted by stop')
        if _search_response_ready(iframe):
            return

        message = _read_visible_payer_alert_text(iframe, skip_informational=True)
        if message and _classify_payer_alert_text(message) == 'transient_wait':
            if not logged_transient:
                log_to_gui("  ℹ️ Payer still retrieving — waiting...\n", "info")
                logged_transient = True
            deadline = min(deadline + 3.0, max_deadline)

        remaining = deadline - time.time()
        if remaining <= 0:
            break
        _interruptible_sleep(min(0.25, remaining))


def submit_search_and_wait(iframe):
    """Click Search and wait for results table or payer not-found alert.

    Wrapped in retry_with_backoff() because the response sometimes
    takes an extra moment to render after a fast network response.
    Returns 'results', 'not_found', 'payer_error', or 'failed'.
    """
    global _last_search_alert_message
    try:
        log_to_gui("  → Submitting search...\n")

        def _submit_once():
            global _last_search_alert_message
            _last_search_alert_message = ''
            iframe.locator(SELECTORS['submit_btn']).click()
            _wait_for_search_response(iframe, timeout_ms=15000)
            outcome = _resolve_search_outcome(iframe)
            if outcome == 'failed':
                raise RuntimeError('Search response did not resolve to table or alert')
            return outcome

        outcome = retry_with_backoff("Submit search", _submit_once, attempts=3)
        if outcome == 'invalid_member_id':
            log_to_gui("  ℹ️ Invalid member / subscriber id\n", "info")
        elif outcome == 'not_found':
            log_to_gui("  ℹ️ Claim not found in adjudication system\n", "info")
        elif outcome == 'payer_error':
            log_to_gui("  ℹ️ Payer returned a search error — skipping row\n", "info")
        elif outcome == 'results':
            log_to_gui("  ✓ Search results loaded\n", "success")
            # Sort is applied once in process_one_row / callers — do not sort here.
            # aria-sort is often empty on Availity, so a second click toggles
            # back to descending and breaks multi-claim row indices.
        return outcome
    except Exception as e:
        log_to_gui(f"  ❌ Search failed: {e}\n", "error")
        return 'failed'


# ============================================================================
# SECTION 10 — CLAIM LIST EXTRACTION
# ============================================================================

def _finalized_dates_look_ascending(iframe):
    """True when #claimsTable Finalized Date values appear in ascending order.

    Used when ``aria-sort`` is missing/empty (common on Availity) so we do not
    click the sort header a second time and toggle back to descending.
    """
    col_idx = _claims_table_finalized_date_col_index(iframe)
    if col_idx is None:
        return False
    rows = _claims_table_body_rows(iframe)
    count = rows.count()
    if count < 2:
        return True

    prev = None
    checked = 0
    for i in range(count):
        if checked >= 8:
            break
        try:
            raw = _read_finalized_date_from_results_row_data_field(rows.nth(i))
            if not raw:
                raw = _read_claims_table_row_cell(rows.nth(i), col_idx)
            if not raw or raw in ('--', 'N/A'):
                continue
            current = normalize_date(raw)
            if not current:
                continue
            try:
                current_ts = pd.to_datetime(current, errors='coerce')
            except Exception:
                continue
            if pd.isna(current_ts):
                continue
            if prev is not None and current_ts < prev:
                return False
            prev = current_ts
            checked += 1
        except Exception:
            continue
    return checked >= 2


def sort_claim_results_by_finalized_date(iframe):
    """Ensure the claim results table is sorted by Finalized Date ascending.

    Availity loads results in descending order. Returning from claim detail
    resets that order, so row indices must be read only after ascending sort.
    Clicks the header only when not already ascending. ``aria-sort`` is often
    empty on Availity, so date order is used as a fallback guard (a second
    click would toggle back to descending).
    """
    try:
        header = iframe.locator(SELECTORS['claims_finalized_date_header']).first
        header.wait_for(state='visible', timeout=10000)
        sort_attr = (header.get_attribute('aria-sort') or '').lower()
        dates_ascending = False
        if sort_attr != 'ascending':
            dates_ascending = _finalized_dates_look_ascending(iframe)
        already_ascending = sort_attr == 'ascending' or dates_ascending
        if already_ascending:
            return True

        log_to_gui("  → Sorting results by Finalized Date (ascending)...\n")

        def _click_sort_header():
            header.click(timeout=10000)
            iframe.locator('#claimsTable tbody tr').first.wait_for(
                state='visible', timeout=10000
            )

        retry_with_backoff("Sort by Finalized Date", _click_sort_header, attempts=2)
        log_to_gui("  ✓ Results sorted ascending\n", "success")
        return True
    except Exception as e:
        log_to_gui(f"  ⚠️ Could not sort results: {e}\n", "error")
        return False


def _claims_table_finalized_date_col_index(iframe):
    """Return 0-based #claimsTable column index for Finalized Date, or None."""
    try:
        header = iframe.locator(SELECTORS['claims_finalized_date_header']).first
        if header.count() > 0:
            idx = header.evaluate(
                """el => {
                    const row = el.closest('tr');
                    if (!row) return -1;
                    const headers = Array.from(
                        row.querySelectorAll('th[role="columnheader"], th')
                    );
                    return headers.indexOf(el);
                }"""
            )
            if isinstance(idx, int) and idx >= 0:
                return idx
    except Exception:
        pass

    for header_sel in (
        '#claimsTable thead tr[role="row"] th[role="columnheader"]',
        '#claimsTable thead th[role="columnheader"]',
        '#claimsTable thead th',
    ):
        try:
            headers = iframe.locator(header_sel)
            for i in range(headers.count()):
                label = (headers.nth(i).inner_text(timeout=2000) or '').strip().lower()
                if 'finalized' in label:
                    return i
        except Exception:
            continue
    return None


def _read_claims_table_row_cell(row_locator, col_idx):
    """Read visible text from a results-table body cell (td or gridcell)."""
    for cell_sel in ('td', '[role="gridcell"]'):
        try:
            cells = row_locator.locator(cell_sel)
            if cells.count() > col_idx:
                raw = (cells.nth(col_idx).inner_text(timeout=3000) or '').strip()
                if raw:
                    return raw
        except Exception:
            continue
    return ''


def _read_finalized_date_from_results_row_data_field(row_locator):
    """MUI DataGrid rows may expose finalizedDate via data-field on cells."""
    for field in ('finalizedDate', 'finalized_date'):
        try:
            cell = row_locator.locator(f'[data-field="{field}"]')
            if cell.count() > 0:
                raw = (cell.first.inner_text(timeout=3000) or '').strip()
                if raw and raw not in ('--', 'N/A'):
                    return raw
        except Exception:
            continue
    return ''


def read_finalized_date_from_results_row(iframe, row_locator):
    """Read and normalize Finalized Date from one #claimsTable tbody row."""
    try:
        raw = _read_finalized_date_from_results_row_data_field(row_locator)
        if not raw:
            col_idx = _claims_table_finalized_date_col_index(iframe)
            if col_idx is None:
                log_to_gui(
                    "    ⚠️ Finalized Date column not found in results table\n",
                    "error",
                )
                return None
            raw = _read_claims_table_row_cell(row_locator, col_idx)
        if not raw or raw in ('--', 'N/A'):
            log_to_gui(
                "    ⚠️ Finalized Date empty in results table row\n",
                "error",
            )
            return None
        normalized = normalize_date(raw)
        log_to_gui(f"    • Results Finalized Date: {normalized}\n", "info")
        return normalized
    except Exception as e:
        log_to_gui(f"    ⚠️ Could not read results Finalized Date: {e}\n", "error")
        return None


def _peek_results_row_finalized_date(iframe, row_locator):
    """Quietly read Finalized Date from a results row (no GUI log)."""
    try:
        raw = _read_finalized_date_from_results_row_data_field(row_locator)
        if not raw:
            col_idx = _claims_table_finalized_date_col_index(iframe)
            if col_idx is None:
                return None
            raw = _read_claims_table_row_cell(row_locator, col_idx)
        if not raw or raw in ('--', 'N/A'):
            return None
        return normalize_date(raw)
    except Exception:
        return None


def _claims_table_body_rows(iframe):
    """Locator for data rows in the claim search results table."""
    # Prefer direct tbody children so nested tables are not double-counted.
    for sel in (
        '#claimsTable > tbody > tr[role="row"]',
        '#claimsTable > tbody > tr',
        '#claimsTable tbody tr[role="row"]',
        '#claimsTable tbody tr',
    ):
        rows = iframe.locator(sel)
        if rows.count() > 0:
            return rows
    return iframe.locator('#claimsTable tbody tr')


def _claims_table_service_dates_col_index(iframe):
    """Return 0-based #claimsTable column index for Service Dates, or None."""
    for header_sel in (
        '#claimsTable thead tr[role="row"] th[role="columnheader"]',
        '#claimsTable thead th[role="columnheader"]',
        '#claimsTable thead th',
    ):
        try:
            headers = iframe.locator(header_sel)
            for i in range(headers.count()):
                label = (headers.nth(i).inner_text(timeout=2000) or '').strip().lower()
                if 'service date' in label:
                    return i
        except Exception:
            continue
    return None


def _results_date_range_from_row(row_locator, service_dates_col):
    """Return normalised 'MM/DD/YYYY-MM/DD/YYYY' from a results-table row."""
    try:
        date_cell = row_locator.locator('td').nth(service_dates_col)
        cell_text = date_cell.inner_text(timeout=3000).strip()
        if cell_text:
            parsed = parse_service_date_range(cell_text)
            if parsed:
                return parsed

        paragraphs = date_cell.locator('p')
        p_count = paragraphs.count()
        if p_count >= 2:
            from_d = normalize_date(
                paragraphs.nth(0).text_content(timeout=3000).strip()
            )
            to_d = normalize_date(
                paragraphs.nth(1).text_content(timeout=3000).strip()
            )
            return f"{from_d}-{to_d}"
        if p_count == 1:
            single_text = paragraphs.nth(0).text_content(timeout=3000).strip()
            parsed = parse_service_date_range(single_text)
            if parsed:
                return parsed
            single = normalize_date(single_text)
            return f"{single}-{single}"
    except Exception:
        pass
    return None


def _should_filter_results_by_visit_date(payer_config):
    """True when results rows should be narrowed to the spreadsheet VisitDate."""
    cfg = resolve_line_level_config(payer_config)
    explicit = cfg.get('filter_results_by_visit_date')
    if explicit is not None:
        return bool(explicit)
    return bool(cfg.get('fallback_within_range'))


def _narrow_claim_indices_by_visit_date(iframe, indices, visit_date, payer_config):
    """Keep only #claimsTable rows whose Service Dates contain VisitDate."""
    if not indices or not _should_filter_results_by_visit_date(payer_config):
        return indices

    visit_date = coerce_visit_date(visit_date)
    if not visit_date:
        return indices

    col_idx = _claims_table_service_dates_col_index(iframe)
    if col_idx is None:
        log_to_gui(
            "  ⚠️ Cannot filter results by VisitDate — Service Dates column not found\n",
            "error",
        )
        return indices

    cfg = resolve_line_level_config(payer_config)
    target = normalize_date_range(visit_date)
    rows = _claims_table_body_rows(iframe)
    narrowed = []

    for i in indices:
        row_range = _results_date_range_from_row(rows.nth(i), col_idx)
        if not row_range:
            continue
        row_range_norm = normalize_date_range(row_range)
        if visit_date_matches_line(target, row_range_norm, 'exact'):
            narrowed.append(i)
            log_to_gui(
                f"  ✓ Results row {i + 1} exact VisitDate match: {row_range_norm}\n",
                "success",
            )
        elif cfg.get('fallback_within_range') and visit_date_matches_line(
            target, row_range_norm, 'within_range'
        ):
            narrowed.append(i)
            log_to_gui(
                f"  ✓ Results row {i + 1} VisitDate within service range: "
                f"{row_range_norm}\n",
                "success",
            )

    if not narrowed:
        scraped = []
        for i in indices:
            row_range = _results_date_range_from_row(rows.nth(i), col_idx)
            if row_range:
                scraped.append(normalize_date_range(row_range))
        log_to_gui(
            f"  ⚠️ No results row contains VisitDate {target} "
            f"({len(indices)} invoice match(es); scraped: {', '.join(scraped)})\n",
            "error",
        )
    elif len(narrowed) < len(indices):
        log_to_gui(
            f"  ℹ️ Narrowed {len(indices)} invoice match(es) to {len(narrowed)} "
            f"by VisitDate {target}\n",
            "info",
        )
    return narrowed


def _service_date_range_matches(target_range, row_range):
    """True when results Service Dates equal the aggregated bundle range."""
    if not target_range or not row_range:
        return False
    target = normalize_date_range(target_range)
    row = normalize_date_range(row_range)
    return target == row


def _find_claim_row_indices_by_invoice(iframe, invoice_number):
    """0-based indices of results rows whose text contains invoice_number."""
    indices = []
    rows = _claims_table_body_rows(iframe)
    for i in range(rows.count()):
        try:
            text = rows.nth(i).inner_text(timeout=3000) or ''
            if invoice_number in text:
                indices.append(i)
        except Exception:
            continue
    return indices


def _find_claim_row_indices_by_service_dates(iframe, target_range):
    """0-based indices of results rows whose Service Dates match target_range."""
    col_idx = _claims_table_service_dates_col_index(iframe)
    if col_idx is None:
        log_to_gui(
            "    ⚠️ Service Dates column not found in results table\n",
            "error",
        )
        return []

    indices = []
    rows = _claims_table_body_rows(iframe)
    for i in range(rows.count()):
        row = rows.nth(i)
        row_range = _results_date_range_from_row(row, col_idx)
        if _service_date_range_matches(target_range, row_range):
            indices.append(i)
            log_to_gui(
                f"    ✓ Bundle match row {i + 1}: {row_range}\n",
                "success",
            )
    return indices


def _claim_row_locator_at_index(iframe, index):
    """Single results-table row locator by 0-based index."""
    return _claims_table_body_rows(iframe).nth(index)


def resolve_matching_claim_indices(iframe, df, invoice_number,
                                   visit_date=None, payer_config=None):
    """Find results rows by invoice text, else by aggregated Service Dates.

    When ``filter_results_by_visit_date`` is enabled (default for payers with
    ``fallback_within_range``), invoice matches are narrowed to rows whose
    Service Dates contain the spreadsheet VisitDate.

    Returns (indices, match_mode) where match_mode is 'invoice',
    'service_dates', or '' when nothing matched.
    """
    try:
        if iframe.locator('#claimsTable tbody').count() == 0:
            raise Exception("Results table body not found")

        indices = _find_claim_row_indices_by_invoice(iframe, invoice_number)
        if indices:
            log_to_gui(
                f"  ✓ Found {len(indices)} claim(s) by invoice\n",
                "success",
            )
            indices = _narrow_claim_indices_by_visit_date(
                iframe, indices, visit_date, payer_config,
            )
            if indices:
                return indices, 'invoice'
            return [], ''

        target_range = aggregate_invoice_visit_date_range(df, invoice_number)
        if not target_range:
            log_to_gui(
                f"  ⚠️ Invoice {invoice_number} not found; "
                "no VisitDate rows to match bundle range\n",
                "error",
            )
            return [], ''

        visit_rows = sum(
            1
            for _, row in df.iterrows()
            if format_invoice_number(row.get('InvoiceNumber', ''), '')
            == format_invoice_number(invoice_number)
        )
        log_to_gui(
            f"  → Bundle claim fallback: Service Dates {target_range} "
            f"({visit_rows} collection row(s))\n",
            "info",
        )
        indices = _find_claim_row_indices_by_service_dates(iframe, target_range)
        if indices:
            log_to_gui(
                f"  ✓ Found {len(indices)} claim(s) by service dates\n",
                "success",
            )
            return indices, 'service_dates'

        log_to_gui(
            f"  ⚠️ No results row matches bundle range {target_range}\n",
            "error",
        )
        return [], ''
    except Exception as e:
        log_to_gui(f"  ⚠️ {e}\n", "error")
        return [], ''


def find_matching_claims(iframe, invoice_number):
    """Locate all result rows whose visible text contains invoice_number.

    Returns (locator, count).  count == 0 means the invoice was not found.
    Prefer resolve_matching_claim_indices() for bundle fallback.
    """
    try:
        if iframe.locator('#claimsTable tbody').count() == 0:
            raise Exception("Results table body not found")

        indices = _find_claim_row_indices_by_invoice(iframe, invoice_number)
        if not indices:
            raise Exception(f"Invoice {invoice_number} not found in results")

        rows = _claims_table_body_rows(iframe)
        count = len(indices)
        log_to_gui(f"  ✓ Found {count} claim(s)\n", "success")
        return rows, count
    except Exception as e:
        log_to_gui(f"  ⚠️ {e}\n", "error")
        return None, 0


# ============================================================================
# SECTION 11 — CLAIM DETAIL EXTRACTION
# ============================================================================

def _read_selector(iframe, selector, label, timeout=5000, *, quiet=False):
    """Read visible text from a single selector.  Returns '--' if absent.

    Prefixed with _ because it is an implementation detail used only by
    the extraction functions in this section.
    """
    try:
        return iframe.locator(selector).text_content(timeout=timeout).strip()
    except Exception:
        if not quiet:
            log_to_gui(f"    ⚠️ {label} not found\n", "error")
        return '--'


def _visible_mui_date_group(iframe, picker_label):
    """Return the first visible MUI date picker group for the given accessible name."""
    candidates = iframe.get_by_role('group', name=picker_label)
    try:
        count = candidates.count()
    except Exception:
        count = 0
    for i in range(count):
        group = candidates.nth(i)
        try:
            if group.is_visible():
                return group
        except Exception:
            continue
    return _mui_date_group_by_legend(iframe, picker_label)


def _mui_date_group_by_legend(iframe, legend_text):
    """Resolve a date picker group scoped to its fieldset legend (From vs To)."""
    for root_sel in (
        'div.MuiPickersInputBase-root[role="group"]',
        '[role="group"].MuiPickersInputBase-root',
    ):
        group = iframe.locator(
            f'fieldset:has(legend:has-text("{legend_text}")) {root_sel}'
        ).first
        try:
            if group.count() > 0 and group.is_visible():
                return group
        except Exception:
            continue
    group = iframe.locator(
        f'fieldset:has(legend:has-text("{legend_text}")) '
        'div.MuiPickersInputBase-root[role="group"]'
    ).first
    group.wait_for(state='visible', timeout=5000)
    return group


def _resolve_mui_date_group(iframe, picker_label):
    """Resolve picker group: accessible role first, then fieldset legend."""
    try:
        group = _visible_mui_date_group(iframe, picker_label)
        group.wait_for(state='visible', timeout=5000)
        return group
    except Exception:
        return _mui_date_group_by_legend(iframe, picker_label)


def _commit_mui_date_group(page, iframe):
    """Blur the active date picker so React commits before touching another picker."""
    member = iframe.locator(SELECTORS['member_id'])
    member.click()
    member.wait_for(state='visible', timeout=3000)


def _normalize_mui_date_str(date_str):
    """Normalize and validate a date string as MM/DD/YYYY for MUI pickers."""
    parts = split_date_parts(date_str)
    if parts:
        date_str = f'{parts[0]}/{parts[1]}/{parts[2]}'
    if not re.match(r'^\d{2}/\d{2}/\d{4}$', date_str):
        raise ValueError(f'Invalid date format. Expected MM/DD/YYYY, got: {date_str}')
    return date_str


def _focus_mui_date_group_year(page, group):
    """Click the Year spinbutton to isolate focus in this picker group."""
    group.get_by_role('spinbutton', name='Year').click()
    time.sleep(0.05)


def _wait_for_patient_dob_picker(iframe, timeout=15000):
    """Wait until the Patient DOB picker group is visible."""
    _visible_mui_date_group(iframe, 'Patient Date of Birth').wait_for(
        state='visible', timeout=timeout
    )


def _mui_date_group_matches(group, date_str):
    """True when Month/Day/Year spinbuttons match the intended date."""
    month, day, year = date_str.split('/')
    for label, want in [('Month', month), ('Day', day), ('Year', year)]:
        try:
            shown = group.get_by_role('spinbutton', name=label).inner_text(timeout=2000).strip()
            if not shown.isdigit() or int(shown) != int(want):
                return False
        except Exception:
            return False
    return True


_MUI_DATE_TYPE_DELAY = 80


def _read_mui_date_group_value(group):
    """Read visible Month/Day/Year segments as MM/DD/YYYY, or None if unreadable."""
    parts = []
    for label in ('Month', 'Day', 'Year'):
        try:
            parts.append(
                group.get_by_role('spinbutton', name=label).inner_text(timeout=2000).strip()
            )
        except Exception:
            return None
    if len(parts) == 3:
        return f'{parts[0]}/{parts[1]}/{parts[2]}'
    return None


def _mui_date_group_is_empty(group):
    """True when Month/Day/Year segments are blank or the MM/DD/YYYY placeholder."""
    value = _read_mui_date_group_value(group)
    if not value:
        return True
    parts = value.split('/')
    if len(parts) != 3:
        return True
    return not all(part.isdigit() for part in parts)


def _clear_mui_date_section(page, section, name):
    """Backspace-clear one spinbutton section (Year, Day, or Month)."""
    width = _MUI_SECTION_WIDTHS.get(name, 4)
    section.click()
    time.sleep(0.05)
    for _ in range(width):
        page.keyboard.press('Backspace')
        time.sleep(0.02)


def _reset_mui_date_group(page, iframe, group):
    """Clear all spinbutton segments in a MUI date picker group."""
    for name in ('Year', 'Day', 'Month'):
        section = group.get_by_role('spinbutton', name=name)
        _clear_mui_date_section(page, section, name)
    _commit_mui_date_group(page, iframe)


def _service_date_pickers_need_reset(from_group, to_group, from_date, to_date):
    """True when coupled pickers still hold values from a prior search attempt."""
    if _mui_date_group_is_empty(from_group) and _mui_date_group_is_empty(to_group):
        return False
    from_shown = _read_mui_date_group_value(from_group)
    to_shown = _read_mui_date_group_value(to_group)
    return from_shown != from_date or to_shown != to_date


def _reset_stale_service_date_pickers(page, iframe, from_group, to_group):
    """Clear stale Service To / From pickers before refill (To first)."""
    log_to_gui('  → Clearing stale service dates before refill...\n', 'info')
    if not _mui_date_group_is_empty(to_group):
        _reset_mui_date_group(page, iframe, to_group)
    if not _mui_date_group_is_empty(from_group):
        _reset_mui_date_group(page, iframe, from_group)
    _commit_mui_date_group(page, iframe)


def _hidden_mui_date_input(group):
    """Return the hidden backing input for a MUI date picker group."""
    hidden = group.locator('input[aria-hidden="true"]')
    if hidden.count() == 0:
        hidden = group.locator('input.MuiPickersInputBase-input')
    return hidden.first


def _validate_mui_date_filled(group, date_str):
    """Confirm spinbuttons or hidden input reflect the intended date."""
    if _mui_date_group_matches(group, date_str):
        return
    expect(_hidden_mui_date_input(group)).to_have_value(date_str, timeout=3000)


def _wait_for_mui_date_group_match(group, date_str, timeout_ms=1500):
    """Poll spinbuttons until the intended date is visible or timeout."""
    deadline = time.time() + (timeout_ms / 1000)
    while time.time() < deadline:
        if _mui_date_group_matches(group, date_str):
            return True
        time.sleep(0.05)
    return _mui_date_group_matches(group, date_str)


_MUI_DATE_SEGMENT_ORDER = ('Year', 'Day', 'Month')


def _section_value_matches(section, expected):
    """True when the spinbutton section's visible text equals expected (numeric)."""
    try:
        shown = section.inner_text(timeout=1500).strip()
        return shown.isdigit() and int(shown) == int(expected)
    except Exception:
        return False


_MUI_SECTION_WIDTHS = {'Year': 4, 'Day': 2, 'Month': 2}


def _clear_and_fill_section(page, section, name, value, delay=_MUI_DATE_TYPE_DELAY):
    """Clear a pre-populated section with Backspace, then type fresh digits.

    Why Backspace-clear:
    - When a section holds a value (e.g. To's Year = today's year 2026), MUI
      enters replace-mode on the first keystroke and frequently drops the
      remaining digits, leaving the section showing just the first digit
      zero-padded (e.g. typing 2026 over 2026 lands on 0002).
    - Backspacing the section's width clears its value cleanly. Typing then
      fills the empty section digit-by-digit via MUI's shift behavior, the
      same path that works for empty pickers like Patient DOB.
    """
    _clear_mui_date_section(page, section, name)
    section.click()
    time.sleep(0.05)
    page.keyboard.type(value, delay=delay)
    time.sleep(0.08)


def _fill_mui_date_segments(page, group, month, day, year, delay=_MUI_DATE_TYPE_DELAY):
    """Fill Year → Day → Month, clearing each section with Backspace first."""
    year_section = group.get_by_role('spinbutton', name='Year')
    day_section = group.get_by_role('spinbutton', name='Day')
    month_section = group.get_by_role('spinbutton', name='Month')

    _clear_and_fill_section(page, year_section, 'Year', year, delay=delay)
    _clear_and_fill_section(page, day_section, 'Day', day, delay=delay)
    _clear_and_fill_section(page, month_section, 'Month', month, delay=delay)

    # Per-section verify+retry for any section that still doesn't match.
    for section, value, name in (
        (year_section, year, 'Year'),
        (day_section, day, 'Day'),
        (month_section, month, 'Month'),
    ):
        if _section_value_matches(section, value):
            continue
        for _ in range(2):
            _clear_and_fill_section(page, section, name, value, delay=delay + 20)
            if _section_value_matches(section, value):
                break


def _fill_mui_date_continuous(page, group, month, day, year, delay=_MUI_DATE_TYPE_DELAY):
    """Retry variant: same Backspace-clear flow with longer per-key delay."""
    _fill_mui_date_segments(page, group, month, day, year, delay=delay + 30)


def _fill_mui_date_group_year_only(page, group, year):
    """Set only the Year spinbutton (used before cross-year range completion)."""
    year_section = group.get_by_role('spinbutton', name='Year')
    _clear_and_fill_section(page, year_section, 'Year', year)


def _fill_mui_date_group_via_input(page, group, date_str, *, validate_after=True):
    """Fill a resolved picker group via its hidden backing input."""
    date_str = _normalize_mui_date_str(date_str)
    hidden = _hidden_mui_date_input(group)
    hidden.wait_for(state='attached', timeout=5000)
    hidden.click(force=True)
    time.sleep(0.05)
    page.keyboard.press('Control+a')
    page.keyboard.press('Delete')
    time.sleep(0.05)
    page.keyboard.type(date_str, delay=_MUI_DATE_TYPE_DELAY)
    hidden.evaluate(
        """(el, val) => {
            el.value = val;
            el.dispatchEvent(new InputEvent('input', { bubbles: true }));
            el.dispatchEvent(new Event('change', { bubbles: true }));
        }""",
        date_str,
    )
    page.keyboard.press('Tab')
    _wait_for_mui_date_group_match(group, date_str)
    if validate_after:
        _validate_mui_date_filled(group, date_str)


def _fill_mui_date_group_robust(
    page,
    iframe,
    group,
    date_str,
    *,
    commit=True,
    commit_mode='blur',
    retries=2,
    validate_after=True,
):
    """Fill a pre-resolved MUI date group with escalating strategies."""
    date_str = _normalize_mui_date_str(date_str)
    month, day, year = date_str.split('/')
    strategies = ('segments', 'continuous', 'via_input')
    last_error = None

    for attempt in range(retries + 1):
        strategy = strategies[min(attempt, len(strategies) - 1)]
        try:
            if strategy == 'via_input':
                _fill_mui_date_group_via_input(
                    page, group, date_str, validate_after=validate_after,
                )
                return

            if strategy == 'segments':
                _fill_mui_date_segments(page, group, month, day, year)
            else:
                _fill_mui_date_continuous(page, group, month, day, year)

            if commit:
                if commit_mode == 'blur':
                    _commit_mui_date_group(page, iframe)
                else:
                    page.keyboard.press('Tab')
                    _wait_for_mui_date_group_match(group, date_str)

            if validate_after:
                _validate_mui_date_filled(group, date_str)
            return
        except Exception as e:
            last_error = e
            if attempt >= retries:
                break
            shown = _read_mui_date_group_value(group)
            detail = f' (shown: {shown})' if shown else ''
            log_to_gui(
                f'  ⚠️ Date group {strategy} attempt {attempt + 1} failed{detail}, '
                f'retrying...\n',
                'error',
            )

    shown = _read_mui_date_group_value(group)
    detail = f' (shown: {shown})' if shown else ''
    raise RuntimeError(
        f'Failed to fill date group with {date_str}{detail}: {last_error}'
    )


def fill_mui_date_picker(
    page,
    iframe,
    picker_label,
    date_str,
    retries=2,
    validate_after=True,
    commit_mode='tab',
    group=None,
):
    """Fill a MUI X DatePicker via Year/Day/Month spinbuttons.

    Retries: segment-by-segment, then YYYYMMDD continuous, then hidden-input fallback.
    commit_mode: 'tab' for single pickers (DOB), 'blur' to avoid focus bleed in ranges.
    """
    date_str = _normalize_mui_date_str(date_str)
    log_to_gui(f'  → {picker_label}: {date_str}\n', 'info')

    page.bring_to_front()
    resolved_group = group or _resolve_mui_date_group(iframe, picker_label)
    resolved_group.wait_for(state='visible', timeout=5000)
    resolved_group.scroll_into_view_if_needed()

    try:
        _fill_mui_date_group_robust(
            page,
            iframe,
            resolved_group,
            date_str,
            commit=True,
            commit_mode=commit_mode,
            retries=retries,
            validate_after=validate_after,
        )
    except RuntimeError as e:
        raise RuntimeError(
            f'Failed to fill {picker_label} with {date_str}: {e}'
        ) from e


def fill_mui_date_picker_via_input(
    page,
    iframe,
    picker_label,
    date_str,
    retries=2,
    validate_after=True,
    log_header=True,
):
    """Last-resort fill via the MUI hidden input inside the picker group.

    The input id is dynamic (e.g. :r2b:, :r2e:) — locate via group label, not by id.
    """
    parts = split_date_parts(date_str)
    if parts:
        date_str = f'{parts[0]}/{parts[1]}/{parts[2]}'
    if not re.match(r'^\d{2}/\d{2}/\d{4}$', date_str):
        raise ValueError(f'Invalid date format. Expected MM/DD/YYYY, got: {date_str}')

    if log_header:
        log_to_gui(f'  → {picker_label}: {date_str} (input fallback)\n', 'info')
    last_error = None
    group = None
    for attempt in range(retries + 1):
        try:
            page.bring_to_front()
            group = _visible_mui_date_group(iframe, picker_label)
            group.wait_for(state='visible', timeout=5000)
            group.scroll_into_view_if_needed()
            _fill_mui_date_group_via_input(
                page, group, date_str, validate_after=validate_after,
            )
            return
        except Exception as e:
            last_error = e
            if attempt >= retries:
                break
            log_to_gui(
                f'  ⚠️ DatePicker input fill attempt {attempt + 1} failed, retrying...\n',
                'error',
            )
            time.sleep(0.5)

    shown = _read_mui_date_group_value(group) if group is not None else None
    detail = f' (shown: {shown})' if shown else ''
    raise RuntimeError(
        f'Failed to fill {picker_label} via input with {date_str}{detail}: {last_error}'
    )


def fill_service_date_range(
    page,
    iframe,
    from_date,
    to_date,
    from_label=None,
    to_label=None,
):
    """Fill coupled MUI service-date range pickers with adaptive cross-year plans."""
    from_label = from_label or SELECTORS['service_from_date_label']
    to_label = to_label or SELECTORS['service_to_date_label']
    from_date = _normalize_mui_date_str(from_date)
    to_date = _normalize_mui_date_str(to_date)
    cross = service_dates_cross_year(from_date, to_date)

    log_to_gui(
        f'  → Service dates: {from_date} → {to_date}'
        f'{" (cross-year)" if cross else ""}\n',
        'info',
    )
    page.bring_to_front()

    from_group = _resolve_mui_date_group(iframe, from_label)
    to_group = _resolve_mui_date_group(iframe, to_label)
    to_group.scroll_into_view_if_needed()

    if _service_date_pickers_need_reset(from_group, to_group, from_date, to_date):
        _reset_stale_service_date_pickers(page, iframe, from_group, to_group)

    plans = service_date_fill_plans(from_date, to_date)
    for plan_idx, plan in enumerate(plans):
        order_label = ' then '.join(plan['order'])
        log_to_gui(
            f'  → Service date plan {plan_idx + 1}/{len(plans)}: '
            f'{plan["mode"]} ({order_label})\n',
            'info',
        )
        try:
            _execute_service_date_fill_plan(
                page, iframe, from_group, to_group, from_date, to_date, plan,
            )
        except Exception as e:
            log_to_gui(f'  ⚠️ Service date plan error: {e}\n', 'error')

        if _service_date_range_valid(from_group, to_group, from_date, to_date):
            log_to_gui(
                f'  ✓ Service dates set — From: '
                f'{_read_mui_date_group_value(from_group)} | To: '
                f'{_read_mui_date_group_value(to_group)} '
                f'(plan {plan_idx + 1})\n',
                'success',
            )
            return

        from_shown = _read_mui_date_group_value(from_group)
        to_shown = _read_mui_date_group_value(to_group)
        log_to_gui(
            f'  ⚠️ Service date read-back: From {from_shown} (want {from_date}), '
            f'To {to_shown} (want {to_date}) — trying next plan...\n',
            'error',
        )

    from_shown = _read_mui_date_group_value(from_group)
    to_shown = _read_mui_date_group_value(to_group)
    raise RuntimeError(
        f'Service From read-back: {from_shown} (want {from_date}); '
        f'Service To read-back: {to_shown} (want {to_date})'
    )


def _service_date_range_valid(from_group, to_group, from_date, to_date):
    """True when both service-date picker groups match the intended range."""
    return (
        _mui_date_group_matches(from_group, from_date)
        and _mui_date_group_matches(to_group, to_date)
    )


def _service_date_fields_ordered(order, from_group, to_group, from_date, to_date):
    """Return (group, date_str) pairs in the requested fill order."""
    field_map = {
        'from': (from_group, from_date),
        'to': (to_group, to_date),
    }
    return [field_map[name] for name in order]


def _execute_service_date_fill_plan(
    page, iframe, from_group, to_group, from_date, to_date, plan,
):
    """Run one service-date fill plan (robust segments, year-first, or hidden input)."""
    mode = plan['mode']
    fields = _service_date_fields_ordered(
        plan['order'], from_group, to_group, from_date, to_date,
    )

    if mode == 'year_first':
        for group, date_str in fields:
            year = date_str.split('/')[2]
            _focus_mui_date_group_year(page, group)
            _fill_mui_date_group_year_only(page, group, year)
        _commit_mui_date_group(page, iframe)
        for group, date_str in fields:
            _fill_mui_date_group_robust(
                page, iframe, group, date_str, commit_mode='blur',
            )
        return

    if mode == 'robust':
        for group, date_str in fields:
            _focus_mui_date_group_year(page, group)
            _fill_mui_date_group_robust(
                page, iframe, group, date_str, commit_mode='blur',
            )
        return

    if mode == 'hidden_input':
        for group, date_str in fields:
            _fill_mui_date_group_via_input(page, group, date_str)
            _commit_mui_date_group(page, iframe)
        return

    raise ValueError(f'Unknown service date fill plan mode: {mode!r}')


def fill_date_range(
    page,
    iframe,
    from_date,
    to_date,
    from_label=None,
    to_label=None,
):
    """Fill Service From / To without focus bleed between coupled MUI pickers."""
    fill_service_date_range(
        page, iframe, from_date, to_date, from_label=from_label, to_label=to_label
    )


def _read_first_selector(selectors, iframe, label, timeout=3000):
    """Try each selector in order and return the first non-empty result.

    Used for fields (Billed/Paid Amount) where Availity uses different
    attribute casing across environments — we try all known variants.
    """
    for sel in selectors:
        try:
            val = iframe.locator(sel).text_content(timeout=timeout).strip()
            if val:
                return val
        except Exception:
            continue
    log_to_gui(f"    ⚠️ {label} not found (tried {len(selectors)} selectors)\n", "error")
    return '--'


def extract_claim_header(iframe, payer_config=None):
    """Scrape the summary panels on the open claim detail page.

    Returns a dict whose keys match OUTPUT_COLUMNS.
    """
    sel = get_selector_profile(payer_config)
    skip_detail_finalized = bool(
        payer_config and payer_config.get('uses_results_finalized_date')
    )
    data = {
        'Claim ID':       _read_selector(iframe, sel['claim_id_panel'],       'Claim ID'),
        'Claim Status':   (
            _read_first_selector(sel['claim_status_selectors'], iframe, 'Claim Status')
            if sel.get('claim_status_selectors')
            else _read_selector(iframe, sel['claim_status_panel'], 'Claim Status')
        ),
        'Finalized Date': (
            '--'
            if skip_detail_finalized
            else _read_selector(
                iframe, sel['finalized_date_panel'], 'Finalized Date'
            )
        ),
        'Check Number':   _read_selector(iframe, sel['check_number_panel'],   'Check Number'),
        'Check Date':     _read_selector(iframe, sel['check_date_panel'],     'Check Date'),
        'Billed Amount':  _read_first_selector(sel['billed_amount_selectors'], iframe, 'Billed Amount'),
        'Paid Amount':    _read_first_selector(sel['paid_amount_selectors'],   iframe, 'Paid Amount'),
    }
    log_to_gui(f"    • Claim ID: {data['Claim ID']}\n")
    log_to_gui(f"    • Status:   {data['Claim Status']}\n")
    return data


def _line_expand_button_locator(row, sel):
    """Return the expand/toggle button for a line-level service row."""
    if sel.get('line_expand_button'):
        return row.locator(sel['line_expand_button']).first
    return row.locator('td').first.locator('button').first


def _is_line_data_row(row, service_dates_col=None, sel=None):
    """True for service-line data rows (expandable or bundle layout), not detail panels."""
    sel = sel or {}
    try:
        if sel.get('line_expand_button'):
            if row.locator(sel['line_expand_button']).count() > 0:
                return True
        elif row.locator('td').first.locator('button').count() > 0:
            return True
        if service_dates_col is not None:
            return _line_date_range_from_row(row, service_dates_col) is not None
    except Exception:
        pass
    return False


# Default td indices when thead labels cannot be read (legacy Healthfirst layout).
_LINE_TABLE_COL_DEFAULTS = {'service_dates': 3, 'paid': 6, 'billed': 7}


def _resolve_line_table_columns(iframe, payer_config=None):
    """Map Service Dates / Paid / Billed columns from #lineLevelTable thead.

    Uses per-profile ``line_table_col_defaults`` when set (e.g. Aetna).
    Header text wins when present.  Disk cache reuses columns from prior runs.
    """
    sel = get_selector_profile(payer_config)
    payer_name = state.current_payer_name or ''
    output_folder = state.current_batch_output_folder
    cached_cols = None
    if payer_name and output_folder:
        cached_cols = get_line_table_columns(output_folder, payer_name)
    defaults = (
        cached_cols
        if cached_cols
        else sel.get('line_table_col_defaults', _LINE_TABLE_COL_DEFAULTS)
    )
    cols = dict(defaults)
    table_id = sel.get('line_table', SELECTORS['line_table'])
    try:
        headers = iframe.locator(f'{table_id} thead tr[role="row"] th')
        for i in range(headers.count()):
            label = (headers.nth(i).inner_text(timeout=2000) or '').strip().lower()
            if 'service date' in label:
                cols['service_dates'] = i
            elif 'billed amount' in label or label == 'billed':
                cols['billed'] = i
            elif label == 'paid amount':
                cols['paid'] = i
            elif 'reason/remark' in label or 'remark code' in label:
                cols['remark_codes'] = i
    except Exception:
        pass
    if payer_name and output_folder:
        persist_line_table_columns(output_folder, payer_name, cols)
    return cols


def _line_date_range_from_row(row, service_dates_col):
    """Return normalised 'MM/DD/YYYY-MM/DD/YYYY' for a line table row, or None."""
    try:
        date_cell = row.locator('td').nth(service_dates_col)
        cell_text = date_cell.inner_text(timeout=3000).strip()
        if cell_text:
            parsed = parse_service_date_range(cell_text)
            if parsed:
                return parsed

        paragraphs = date_cell.locator('p')
        p_count = paragraphs.count()
        if p_count >= 2:
            from_d = normalize_date(paragraphs.nth(0).text_content(timeout=3000).strip())
            to_d = normalize_date(paragraphs.nth(1).text_content(timeout=3000).strip())
            return f"{from_d}-{to_d}"
        if p_count == 1:
            single_text = paragraphs.nth(0).text_content(timeout=3000).strip()
            parsed = parse_service_date_range(single_text)
            if parsed:
                return parsed
            single = normalize_date(single_text)
            return f"{single}-{single}"
    except Exception:
        pass
    return None


def _scrape_line_data_rows(iframe, payer_config=None):
    """Return every service-line row with amounts and date_range."""
    line_rows = []
    sel = get_selector_profile(payer_config)
    cols = _resolve_line_table_columns(iframe, payer_config)
    table_sel = sel.get('line_table', SELECTORS['line_table'])
    row_sel = sel.get('line_table_body_row', f'{table_sel} tbody tr[role="row"]')
    iframe.locator(table_sel).wait_for(state='visible', timeout=15000)
    rows = iframe.locator(row_sel)
    for idx in range(rows.count()):
        row = rows.nth(idx)
        if not _is_line_data_row(row, cols['service_dates'], sel):
            continue
        date_range = _line_date_range_from_row(row, cols['service_dates'])
        if not date_range:
            continue
        billed = row.locator('td').nth(cols['billed']).text_content(timeout=3000).strip()
        paid   = row.locator('td').nth(cols['paid']).text_content(timeout=3000).strip()
        line_rows.append({
            'row': row,
            'idx': idx,
            'billed': billed,
            'paid': paid,
            'date_range': date_range,
        })
    return line_rows


def find_all_lines_by_visit_date(iframe, visit_date, payer_config=None):
    """Return every line-level row whose service dates match visit_date.

    Each entry is a dict: row, idx, billed, paid, date_range.
    Matching uses ``LINE_LEVEL_CONFIG`` (and optional per-payer ``line_level``
    overrides).  When VisitDate does not match but the claim has exactly one
    service line and ``fallback_to_sole_line`` is enabled, that line is used.
    """
    matches = []
    try:
        visit_date = coerce_visit_date(visit_date)
        target = normalize_date_range(visit_date)
        cfg = resolve_line_level_config(payer_config)
        all_lines = _scrape_line_data_rows(iframe, payer_config)
        log_to_gui(
            f"    → Searching {len(all_lines)} line(s) for VisitDate: {target}\n"
        )

        filtered = filter_line_rows_by_visit_date(
            all_lines, visit_date, cfg=cfg,
        )
        seen_keys = set()
        for entry in filtered:
            dedupe_key = (entry['date_range'], entry['billed'], entry['paid'])
            if dedupe_key in seen_keys:
                continue
            seen_keys.add(dedupe_key)
            matches.append(entry)
            log_to_gui(
                f"    ✓ Line match #{len(matches)}: {entry['date_range']} "
                f"(Billed: {entry['billed']}, Paid: {entry['paid']})\n",
                "success",
            )

        if not matches:
            log_to_gui("    ⚠️ No matching line found\n", "error")
        elif len(matches) > 1:
            log_to_gui(
                f"    ℹ️ {len(matches)} line(s) for {target} — "
                f"amounts use first match; denials merged into one field\n",
                "info",
            )
    except Exception as e:
        log_to_gui(f"    ⚠️ Line table error: {e}\n", "error")
    return matches


def find_line_by_visit_date(iframe, visit_date, occurrence=0, payer_config=None):
    """Search the line-level table for the Nth row matching visit_date (0-based).

    Both the target date (from CSV) and the scraped dates are normalised to
    'MM/DD/YYYY-MM/DD/YYYY' before comparison to avoid false mismatches.

    Returns (row_locator, row_index, billed_str, paid_str).
    Returns (None, -1, '--', '--') if no match is found.
    """
    matches = find_all_lines_by_visit_date(iframe, visit_date, payer_config)
    if occurrence < len(matches):
        m = matches[occurrence]
        return m['row'], m['idx'], m['billed'], m['paid']
    return None, -1, '--', '--'


# ============================================================================
# SECTION 12 — DENIAL REASON EXTRACTION (INLINE)
# ============================================================================

def collect_line_data_rows(iframe, visit_date=None, payer_config=None):
    """Return service-line rows for denial/amount extraction.

    When ``LINE_LEVEL_CONFIG['filter_by_visit_date']`` is True (default), only
    rows whose service dates match ``visit_date`` are returned.
    """
    visit_date = coerce_visit_date(visit_date)
    try:
        all_lines = _scrape_line_data_rows(iframe, payer_config)
    except Exception as e:
        log_to_gui(f"    ⚠️ Line table error: {e}\n", "error")
        return []

    cfg = resolve_line_level_config(payer_config)
    if not cfg.get('filter_by_visit_date', True):
        return all_lines

    filtered = filter_line_rows_by_visit_date(
        all_lines, visit_date, cfg=cfg,
    )
    if visit_date:
        target = normalize_date_range(visit_date)
        log_to_gui(
            f"    → Line-level filter: {len(filtered)} of {len(all_lines)} "
            f"row(s) match VisitDate {target}\n"
        )
    return filtered


def _format_denial_from_code_pairs(lines_by_date):
    """Format denial text from code/description pairs grouped by service date.

    - One service date, one code:  ``[CODE] - Description``
    - One service date, many codes: ``[CODE] Description, [CODE] Description``
      (same visit date — always one line, never numbered 1/2)
    - Multiple service dates (bundle): one line per code —
      ``MM/DD/YYYY-CODE-Description`` (newline-separated)
    """
    active = {k: v for k, v in lines_by_date.items() if v}
    if not active:
        return '--'

    unique_dates = list(active.keys())
    if len(unique_dates) == 1:
        pairs = active[unique_dates[0]]
        if len(pairs) == 1:
            code, desc = pairs[0]
            return f"[{code}] - {desc}"
        return ', '.join(f"[{code}] {desc}" for code, desc in pairs)

    parts = []
    for date_range in sorted(unique_dates):
        visit = date_range.split('-')[0]
        for code, desc in active[date_range]:
            parts.append(f"{visit}-{code}-{desc}")
    return '\n'.join(parts)


def _extract_code_pairs_from_row(iframe, matching_row, payer_config=None):
    """Expand one service line and return (code, description) pairs."""
    sel = get_selector_profile(payer_config)
    expand_btn = _line_expand_button_locator(matching_row, sel)
    remark_sel = sel.get('remark_codes_grid', SELECTORS['remark_codes_grid'])
    try:
        log_to_gui("    → Expanding line for denial codes...\n")
        expand_btn.click()
        try:
            iframe.locator(remark_sel).first.wait_for(state='visible', timeout=5000)
        except Exception:
            matching_row.locator('xpath=following-sibling::tr[1]').first.wait_for(
                state='visible', timeout=3000,
            )

        remark_text = _read_remark_codes(iframe, matching_row, payer_config)
        if not remark_text:
            return []

        log_to_gui(f"    • Codes: {remark_text}\n")
        return _lookup_code_pairs(iframe, remark_text)

    except Exception as e:
        log_to_gui(f"    ⚠️ Denial extraction error: {e}\n", "error")
        return []
    finally:
        try:
            expand_btn.click()
            try:
                iframe.locator(remark_sel).first.wait_for(state='hidden', timeout=3000)
            except Exception:
                pass
        except Exception:
            pass


def extract_line_level_denial_reason(iframe, claim_status, payer_config=None,
                                    visit_date=None):
    """Build denial text from service lines matching the input VisitDate.

    Only rows whose service dates match ``visit_date`` are expanded (per
    ``LINE_LEVEL_CONFIG``).  Lines sharing a visit date are merged onto one
    line; different dates (bundle claims) each get their own
    ``date-code-description`` line.

    When ``uses_claim_level_remark_fallback`` is set and line-level read finds
    nothing, reads Reason/Remark Code from the claim info panel instead.
    """
    lines_by_date = {}
    for row_info in collect_line_data_rows(iframe, visit_date, payer_config):
        date_range = row_info['date_range']
        lines_by_date.setdefault(date_range, [])

        if not should_extract_denial(
            claim_status, row_info['billed'], row_info['paid'], row_info['row']
        ):
            continue

        for pair in _extract_code_pairs_from_row(
            iframe, row_info['row'], payer_config,
        ):
            if pair not in lines_by_date[date_range]:
                lines_by_date[date_range].append(pair)

    result = _format_denial_from_code_pairs(lines_by_date)
    if result != '--':
        log_to_gui(f"    ✓ Denial reason: {result}\n", "success")
        return result

    if payer_config and payer_config.get('uses_claim_level_remark_fallback'):
        remark_text = _read_claim_level_remark_codes(iframe)
        if remark_text:
            log_to_gui(f"    • Claim-level codes: {remark_text}\n")
            pairs = _lookup_code_pairs(iframe, remark_text)
            if pairs:
                result = _format_denial_from_code_pairs({'_single': pairs})
                log_to_gui(f"    ✓ Denial reason (claim panel): {result}\n", "success")
                return result

    return result


def extract_denial_codes_inline(iframe, matching_row, _row_index, payer_config=None):
    """Expand a line row and format denial for that row only (fallback path)."""
    pairs = _extract_code_pairs_from_row(iframe, matching_row, payer_config)
    if not pairs:
        return '--'
    return _format_denial_from_code_pairs({'_single': pairs})


def _read_claim_level_remark_codes(iframe):
    """Read Reason/Remark Code from the claim detail info panel (SWHNY path).

    Returns stripped code text (e.g. ``39``) or empty string if absent.
    """
    for sel in SELECTORS['reason_remark_code_panel_selectors']:
        try:
            val = iframe.locator(sel).text_content(timeout=3000).strip()
            if val and val not in ('--', 'Reason/Remark Code'):
                log_to_gui(f"    ℹ️ Claim-level remark code: {val}\n", "info")
                return val
        except Exception:
            continue
    return ''


def _read_remark_codes(iframe, matching_row, payer_config=None):
    """Find and return the Reason/Remark Codes text from the expanded inline panel.

    Strategy:
      1. Aetna: read Reason/Remark column from expanded sibling row or same row.
      2. Search the line table for MUI grid cells labeled Reason/Remark Codes;
         the last match in DOM order is the most recently expanded row.
      3. Fallback: walk up to 4 sibling <tr> elements of the expanded row.
      4. Legacy fallback: pre-MUI label + following-sibling <p> layout.
    Returns an empty string if nothing is found.
    """
    sel = get_selector_profile(payer_config)
    value_sel = sel.get('remark_codes_value', SELECTORS['remark_codes_value'])

    def _value_from_grid_cell(cell):
        return cell.locator(value_sel).text_content(timeout=3000).strip()

    try:
        col_defaults = sel.get('line_table_col_defaults', {})
        remark_col = col_defaults.get('remark_codes')
        if remark_col is not None:
            for offset in range(0, 5):
                target_row = (
                    matching_row
                    if offset == 0
                    else matching_row.locator(f'xpath=following-sibling::tr[{offset}]')
                )
                try:
                    text = target_row.locator('td').nth(remark_col).inner_text(
                        timeout=3000,
                    ).strip()
                    if text:
                        return text
                except Exception:
                    continue

        cells = iframe.locator(sel.get('remark_codes_grid', SELECTORS['remark_codes_grid']))
        count = cells.count()
        if count > 0:
            return _value_from_grid_cell(cells.nth(count - 1))

        for offset in range(1, 5):
            sibling = matching_row.locator(f'xpath=following-sibling::tr[{offset}]')
            cell = sibling.locator(
                '.MuiGrid-root:has(p:has-text("Reason/Remark Codes"))'
            )
            if cell.count() > 0:
                return _value_from_grid_cell(cell.first)

        # Legacy layout (font-weight-bold label + adjacent <p>)
        line_table = iframe.locator(sel.get('line_table', SELECTORS['line_table']))
        headers = line_table.locator(
            'p.font-weight-bold:has-text("Reason/Remark Codes")'
        )
        count = headers.count()
        if count > 0:
            return headers.nth(count - 1).locator(
                'xpath=following-sibling::p[1]'
            ).text_content(timeout=3000).strip()

        for offset in range(1, 5):
            sibling = matching_row.locator(f'xpath=following-sibling::tr[{offset}]')
            header = sibling.locator(
                'p.font-weight-bold:has-text("Reason/Remark Codes")'
            )
            if header.count() > 0:
                return header.locator(
                    'xpath=following-sibling::p[1]'
                ).text_content(timeout=3000).strip()

        log_to_gui("    ⚠️ Remark codes panel not found\n", "error")
        return ''
    except Exception as e:
        log_to_gui(f"    ⚠️ Remark codes read error: {e}\n", "error")
        return ''


def _lookup_code_pairs(iframe, remark_codes_text):
    """Resolve remark codes to (code, description) pairs via the codes table."""
    try:
        iframe.locator(SELECTORS['codes_table']).scroll_into_view_if_needed()
        time.sleep(0.5)
    except Exception:
        pass

    pairs = []
    seen_codes = set()
    for code in (c.strip() for c in remark_codes_text.split(',')):
        if not code or code in seen_codes:
            continue
        seen_codes.add(code)
        desc = code
        try:
            row = iframe.locator(
                f'#codesTable tbody tr:has(td:text("Remark")):has(td:text-is("{code}"))'
            )
            if row.count() > 0:
                desc = row.locator('td').nth(2).text_content(timeout=3000).strip() or code
                log_to_gui(f"      • {code}: {desc}\n")
        except Exception:
            pass
        pairs.append((code, desc))
    return pairs


def should_extract_denial(claim_status, billed, paid, matching_row):
    """Decide whether inline denial code extraction is worth attempting.

    Returns False (skip) when:
      - Claim is PENDING — no adjudication data exists yet.
      - No line-level row was found — we have nothing to expand.
      - Claim was fully paid — billed == paid (both non-zero), so no denial.
      - Paid exceeds billed — overpayment, so no denial.
    """
    if claim_status.upper() == 'PENDING':
        log_to_gui("    ℹ️ PENDING — skipping denial codes\n", "info")
        return False
    if matching_row is None:
        return False
    billed_val = parse_amount(billed)
    paid_val   = parse_amount(paid)
    if billed_val is not None and paid_val is not None and billed_val > 0:
        if paid_val == billed_val:
            log_to_gui("    ℹ️ Fully paid — skipping denial codes\n", "info")
            return False
        if paid_val > billed_val:
            log_to_gui("    ℹ️ Overpayment — skipping denial codes\n", "info")
            return False
    return True


# ============================================================================
# SECTION 13 — DENIAL REASON EXTRACTION (REMITTANCE VIEWER)
# ============================================================================

def _has_check_number(claim_data):
    """True when claim detail includes a usable check/EFT number (Villagecaremax gate)."""
    check_number = (claim_data.get('Check Number') or '').strip()
    return bool(check_number) and check_number not in ('--', 'N/A')


def _remit_search_not_found(iframe):
    """True when the remittance viewer shows the no-results message."""
    try:
        return iframe.locator(SELECTORS['remit_not_found']).first.is_visible(timeout=1000)
    except Exception:
        return False


def _wait_remit_search_outcome(iframe, timeout_ms=30000):
    """Wait until remittance search shows results table or not-found message."""
    not_found = iframe.locator(SELECTORS['remit_not_found'])
    table = iframe.locator(SELECTORS['remit_table'])
    combined = not_found.or_(table)
    combined.first.wait_for(state='visible', timeout=timeout_ms)
    return _remit_search_not_found(iframe)


def _open_remit_claim_details(page, claim_id):
    """Open remittance viewer, search by claim_id, and land on claim-details.

    Returns (remit_page, iframe) on success.  Returns (None, None) when the
    claim is not found.  Caller must reset the viewer and close remit_page.
    """
    if not claim_id or claim_id == '--':
        return None, None

    remit_page = page.context.new_page()
    remit_page.goto(REMITTANCE_URLS['search'], wait_until='domcontentloaded', timeout=45000)
    remit_page.wait_for_selector(SELECTORS['iframe'], state='attached', timeout=30000)

    iframe = remit_page.frame_locator(SELECTORS['iframe'])

    claim_tab = iframe.locator(SELECTORS['remit_claim_tab'])
    claim_tab.first.wait_for(state='visible', timeout=20000)
    claim_tab.first.click(timeout=15000)

    search_input = iframe.locator(SELECTORS['remit_search_input'])
    search_input.first.wait_for(state='visible', timeout=20000)
    search_input.first.fill(claim_id)
    iframe.locator(SELECTORS['remit_search_btn']).first.click(timeout=15000)

    if _wait_remit_search_outcome(iframe):
        log_to_gui("    ℹ️ Remittance not found in viewer\n", "info")
        remit_page.goto(REMITTANCE_URLS['home'], wait_until='domcontentloaded', timeout=45000)
        remit_page.close()
        return None, None

    iframe.locator(SELECTORS['remit_table']).first.wait_for(state='visible', timeout=30000)
    claim_link = iframe.locator(f'a[id^="claimNumber"]:has-text("{claim_id}")')
    claim_link.first.wait_for(state='visible', timeout=20000)
    claim_link.first.click(timeout=15000)

    remit_page.wait_for_url("**claim-details**", timeout=45000)
    return remit_page, iframe


def _reset_and_close_remit_page(remit_page):
    """Navigate remittance viewer home and close the tab."""
    if remit_page is None:
        return
    try:
        remit_page.goto(REMITTANCE_URLS['home'], wait_until='domcontentloaded', timeout=45000)
    except Exception:
        pass
    try:
        remit_page.close()
    except Exception:
        pass


def _looks_like_remit_disclaimer(text):
    """True when scraped text is the Molina remittance disclaimer, not a dollar amount."""
    lower = (text or '').lower()
    return (
        'disclaimer' in lower
        or 'remittance viewer reflects' in lower
        or '835 file' in lower
        or 'explanation of payments' in lower
    )


def _normalize_remit_amount_text(text):
    """Strip accounting-style parentheses so ``($0.00)`` parses as zero."""
    cleaned = (text or '').strip()
    if cleaned.startswith('(') and cleaned.endswith(')'):
        inner = cleaned[1:-1].strip()
        if inner and not inner.startswith('$') and inner[0].isdigit():
            return f'${inner}'
        return inner
    return cleaned


def _is_remit_currency_value(text):
    """True when text parses as a currency amount and is not boilerplate."""
    if not text or text.strip() in ('', '--'):
        return False
    if _looks_like_remit_disclaimer(text):
        return False
    return parse_amount(_normalize_remit_amount_text(text)) is not None


# Remittance amount probe tuning — short timeout while searching, longer on cached path.
_REMIT_AMOUNT_PROBE_MS = 400
_REMIT_AMOUNT_CONFIRM_MS = 2000
# Session cache: field_key ('charge'|'payment') -> winning path dict (logged on first hit).
_remit_amount_paths = {}

_REMIT_GRID_LAYOUTS = (
    ('table', 'thead th', 'tbody tr', 'td'),
    ('div[role="table"]', 'div[role="columnheader"]', 'div[role="row"]', 'div[role="cell"]'),
)


def _log_remit_amount_winner(field_key, path, elapsed_ms):
    """Log which remittance amount strategy succeeded (for pinning in config later)."""
    detail = path.get('header') or path.get('label') or path.get('selector', '')[:60]
    log_to_gui(
        f"    ℹ️ Remittance {field_key} via {path['method']}/{path['root']}/{detail} "
        f"({elapsed_ms:.0f}ms)\n",
        "info",
    )


def _read_first_currency_selector(selectors, root, label, timeout=_REMIT_AMOUNT_PROBE_MS,
                                  log_miss=True, selector_offset=0):
    """Like _read_first_selector but only accepts parseable currency values."""
    for idx, sel in enumerate(selectors):
        try:
            val = root.locator(sel).first.text_content(timeout=timeout).strip()
            if _is_remit_currency_value(val):
                return val, {
                    'method': 'xpath',
                    'selector': sel,
                    'selector_idx': selector_offset + idx,
                }
        except Exception:
            continue
    if log_miss:
        log_to_gui(f"    ⚠️ {label} not found (tried {len(selectors)} selectors)\n", "error")
    return '--', None


def _remit_search_roots(iframe):
    """Return ordered (name, locator) search roots for remittance claim-details."""
    roots = []
    try:
        iframe.locator(SELECTORS['remit_adj_table']).first.wait_for(
            state='visible', timeout=30000,
        )
    except Exception:
        pass

    adj = iframe.locator(SELECTORS['remit_adj_table']).first
    for root_name, ancestor_xpath in (
        ('adj_card', 'xpath=ancestor::div[contains(@class,"card")][1]'),
        ('adj_paper', 'xpath=ancestor::div[contains(@class,"MuiPaper")][1]'),
        ('adj_card_body', 'xpath=ancestor::div[contains(@class,"card-body")][1]'),
    ):
        try:
            scope = adj.locator(ancestor_xpath)
            if scope.count() > 0:
                roots.append((root_name, scope.first))
                break
        except Exception:
            continue

    service_table_sel = SELECTORS.get('remit_service_line_table')
    if service_table_sel:
        try:
            table = iframe.locator(service_table_sel)
            if table.count() > 0:
                roots.append(('service_line_table', table.first))
        except Exception:
            pass

    roots.append(('iframe', iframe))
    return roots


def _read_remit_grid_column(root, header_match, header_locator, row_locator, cell_locator,
                            timeout=_REMIT_AMOUNT_PROBE_MS):
    """Read the first currency cell under a matching column header in a grid/table."""
    header_match = (header_match or '').strip().lower()
    if not header_match:
        return '--', None

    try:
        headers = root.locator(header_locator)
        col_idx = None
        for header_idx in range(headers.count()):
            header_text = (headers.nth(header_idx).inner_text(timeout=timeout) or '').strip().lower()
            if header_match in header_text:
                col_idx = header_idx
                break
        if col_idx is None:
            return '--', None

        rows = root.locator(row_locator)
        for row_idx in range(rows.count()):
            cells = rows.nth(row_idx).locator(cell_locator)
            if col_idx >= cells.count():
                continue
            val = (cells.nth(col_idx).text_content(timeout=timeout) or '').strip()
            if _is_remit_currency_value(val):
                return val, {'header': header_match, 'col_idx': col_idx, 'row_idx': row_idx}
    except Exception:
        pass

    return '--', None


def _probe_remit_table_column(root, header_match):
    """Probe HTML table and role=table grids for a column header match."""
    for grid_kind, header_sel, row_sel, cell_sel in _REMIT_GRID_LAYOUTS:
        try:
            tables = root.locator(grid_kind)
            for table_idx in range(tables.count()):
                val, detail = _read_remit_grid_column(
                    tables.nth(table_idx), header_match, header_sel, row_sel, cell_sel,
                )
                if val not in ('', '--'):
                    path = {
                        'method': 'table_column',
                        'grid': grid_kind,
                        'table_idx': table_idx,
                        **detail,
                    }
                    return val, path
        except Exception:
            continue
    return '--', None


def _probe_remit_label_row(root, label_text):
    """Find a label row and return the currency value in that row."""
    try:
        rows = root.locator('div.row, tr').filter(has_text=label_text)
        for row_idx in range(rows.count()):
            row = rows.nth(row_idx)
            row_text = (row.inner_text(timeout=_REMIT_AMOUNT_PROBE_MS) or '').strip().lower()
            if 'disclaimer' in row_text:
                continue
            for value_sel in (
                '[class*="text-right"]',
                'td:last-child',
                'div[role="cell"]:last-child',
            ):
                try:
                    val = (
                        row.locator(value_sel).first.text_content(timeout=_REMIT_AMOUNT_PROBE_MS)
                        or ''
                    ).strip()
                    if _is_remit_currency_value(val):
                        return val, {
                            'method': 'label_row',
                            'label': label_text,
                            'value_sel': value_sel,
                            'row_idx': row_idx,
                        }
                except Exception:
                    continue
    except Exception:
        pass
    return '--', None


def _execute_remit_amount_path(roots_by_name, path):
    """Re-run a cached remittance amount path (fast path for subsequent claims)."""
    root = roots_by_name.get(path['root'])
    if root is None:
        return '--'

    method = path['method']
    if method == 'table_column':
        grid_kind = path['grid']
        layout = next((g for g in _REMIT_GRID_LAYOUTS if g[0] == grid_kind), None)
        if not layout:
            return '--'
        _, header_sel, row_sel, cell_sel = layout
        try:
            table = root.locator(grid_kind).nth(path.get('table_idx', 0))
            val, _ = _read_remit_grid_column(
                table, path['header'], header_sel, row_sel, cell_sel,
                timeout=_REMIT_AMOUNT_CONFIRM_MS,
            )
            return val
        except Exception:
            return '--'

    if method == 'label_row':
        try:
            row = root.locator('div.row, tr').filter(
                has_text=path['label'],
            ).nth(path.get('row_idx', 0))
            val = (
                row.locator(path['value_sel']).first.text_content(
                    timeout=_REMIT_AMOUNT_CONFIRM_MS,
                ) or ''
            ).strip()
            return val if _is_remit_currency_value(val) else '--'
        except Exception:
            return '--'

    if method == 'xpath':
        selectors = SELECTORS.get(path.get('selector_key', ''), [])
        idx = path.get('selector_idx', 0)
        if idx < len(selectors):
            try:
                val = root.locator(selectors[idx]).first.text_content(
                    timeout=_REMIT_AMOUNT_CONFIRM_MS,
                ).strip()
                return val if _is_remit_currency_value(val) else '--'
            except Exception:
                return '--'
    return '--'


def _discover_remit_amount(roots, field_key, header_matches, selector_key, label_rows,
                            pinned_path=None):
    """Find billed/paid amount and cache the winning path for this session."""
    roots_by_name = dict(roots)
    if pinned_path:
        val = _execute_remit_amount_path(roots_by_name, pinned_path)
        if val not in ('', '--'):
            return val

    if field_key in _remit_amount_paths:
        cached = _remit_amount_paths[field_key]
        val = _execute_remit_amount_path(roots_by_name, cached)
        if val not in ('', '--'):
            return val

    if isinstance(header_matches, str):
        header_matches = (header_matches,)
    selectors = SELECTORS.get(selector_key, [])

    for root_name, root in roots:
        for header_match in header_matches:
            val, path = _probe_remit_table_column(root, header_match)
            if val not in ('', '--'):
                path['root'] = root_name
                _remit_amount_paths[field_key] = path
                return val

        for label_text in label_rows:
            val, path = _probe_remit_label_row(root, label_text)
            if val not in ('', '--'):
                path['root'] = root_name
                _remit_amount_paths[field_key] = path
                return val

        val, path = _read_first_currency_selector(
            selectors, root, field_key, log_miss=False,
        )
        if val not in ('', '--') and path:
            path['root'] = root_name
            path['selector_key'] = selector_key
            _remit_amount_paths[field_key] = path
            return val

    log_to_gui(
        f"    ⚠️ Remittance {field_key} not found "
        f"(probed {len(roots)} roots)\n",
        "error",
    )
    return '--'


def _read_remit_line_item_amounts(iframe, payer_config=None):
    """Read Charge Amount (billed) and Payment Amount (paid) in one pass."""
    roots = _remit_search_roots(iframe)
    results = {}
    profile = (payer_config or {}).get('remittance_amount_profile', {})

    for field_key, headers, selector_key, labels in (
        (
            'charge',
            (SELECTORS.get('remit_charge_header_match', 'charge amount'),),
            'remit_charge_amount_selectors',
            ('Charge Amount',),
        ),
        (
            'payment',
            (
                SELECTORS.get('remit_payment_header_match', 'payment amount'),
                'paid amount',
            ),
            'remit_payment_amount_selectors',
            ('Payment Amount', 'Paid Amount'),
        ),
    ):
        t0 = time.time()
        had_cache = field_key in _remit_amount_paths or field_key in profile
        val = _discover_remit_amount(
            roots, field_key, headers, selector_key, labels,
            pinned_path=profile.get(field_key),
        )
        if val not in ('', '--') and field_key in _remit_amount_paths:
            if not had_cache:
                _log_remit_amount_winner(
                    field_key, _remit_amount_paths[field_key], (time.time() - t0) * 1000,
                )
        results[field_key] = val

    payer_name = state.current_payer_name or ''
    output_folder = state.current_batch_output_folder
    if payer_name and output_folder and _remit_amount_paths:
        persist_remit_amount_paths(output_folder, payer_name, _remit_amount_paths)

    return results.get('charge', '--'), results.get('payment', '--')


def extract_remittance_claim_details(page, claim_id, prefetch_denial=False,
                                     payer_config=None):
    """Open remittance viewer once; read amounts and optionally denial."""
    remit_page = None
    result = {'billed': '--', 'paid': '--', 'denial': '--'}
    try:
        remit_page, iframe = _open_remit_claim_details(page, claim_id)
        if remit_page is None:
            result['denial'] = REMIT_NOT_FOUND_DENIAL
            return result

        billed, paid = _read_remit_line_item_amounts(iframe, payer_config=payer_config)
        result['billed'] = billed
        result['paid'] = paid

        if prefetch_denial:
            adj_table = iframe.locator(SELECTORS['remit_adj_table'])
            adj_table.first.wait_for(state='visible', timeout=30000)
            result['denial'] = _read_remit_denial(iframe, adj_table)

        return result
    except Exception as e:
        log_to_gui(f"    ⚠️ Remittance claim-details failed: {e}\n", "error")
        return result
    finally:
        _reset_and_close_remit_page(remit_page)


def extract_amounts_from_remittance(page, claim_id):
    """Open remittance viewer and read Billed and Paid amounts for claim_id."""
    details = extract_remittance_claim_details(page, claim_id, prefetch_denial=False)
    billed, paid = details['billed'], details['paid']
    if billed not in ('', '--'):
        log_to_gui(f"    • Billed (remittance charge): {billed}\n")
    if paid not in ('', '--'):
        log_to_gui(f"    • Paid (remittance payment): {paid}\n")
    return billed, paid


def extract_denial_reason_villagecaremax(page, claim_id):
    """Open the Villagecaremax remittance viewer in a new tab and read denial reason.

    Opens a new tab, searches by claim_id, reads the Adjustments table,
    then closes the tab.  Returns '--' on any failure.
    """
    if not claim_id or claim_id == '--':
        return '--'

    remit_page = None
    try:
        remit_page, iframe = _open_remit_claim_details(page, claim_id)
        if remit_page is None:
            return REMIT_NOT_FOUND_DENIAL

        adj_table = iframe.locator(SELECTORS['remit_adj_table'])
        adj_table.first.wait_for(state='visible', timeout=30000)
        return _read_remit_denial(iframe, adj_table)

    except Exception as e:
        log_to_gui(f"    ⚠️ Villagecaremax remittance failed: {e}\n", "error")
        return '--'
    finally:
        _reset_and_close_remit_page(remit_page)


def _read_remit_denial(iframe, adj_table):
    """Read and clean the denial reason text from the Adjustments table.

    Expands 'View More' buttons before reading to capture full text.
    Falls back to the Claim Adjustment column when the remark column
    only contains the generic 'No remittance advice codes applicable.' message.
    """
    remark_cell = adj_table.locator('tbody tr td').nth(2)
    if remark_cell.count() == 0:
        return '--'

    _expand_view_more(remark_cell)
    denial = remark_cell.inner_text(timeout=10000).strip()
    denial = denial.replace("View More", "").replace("View Less", "").strip()

    if denial != "No remittance advice codes applicable.":
        return denial or '--'

    # Generic remark → fall back to Claim Adjustment column (4th column)
    adj_cell = adj_table.locator('tbody tr td').nth(3)
    if adj_cell.count() == 0:
        return '--'
    _expand_view_more(adj_cell)
    # Data rows use role="cell"; header row uses role="columnheader"
    code_desc = adj_cell.locator(
        'div[role="row"] div[role="cell"].col-sm-4'
    ).nth(1)
    if code_desc.count() == 0:
        code_desc = adj_cell.locator('div[role="row"] div[role="cell"]').nth(1)
    if code_desc.count() == 0:
        return '--'
    text = code_desc.inner_text(timeout=10000).strip()
    text = text.replace("View More", "").replace("View Less", "").strip()
    return text or '--'


def _expand_view_more(container):
    """Click all 'View More' buttons inside container to reveal truncated text."""
    buttons = container.locator('button:has-text("View More")')
    for idx in range(buttons.count()):
        try:
            btn = buttons.nth(idx)
            btn.click(timeout=3000)
            try:
                btn.wait_for(state='hidden', timeout=2000)
            except Exception:
                pass
        except Exception:
            pass


# ============================================================================
# SECTION 14 — PAYER-SPECIFIC CLAIM ENRICHMENT
# ============================================================================
# These three functions encapsulate all payer-branching logic so that
# process_one_claim() is fully payer-agnostic.

def enrich_claim_with_amounts(iframe, claim_data, row_data, payer_config,
                            is_bundle_claim=False, page=None):
    """Populate Billed Amount and Paid Amount based on payer config.

    uses_line_level=True  → find the matching service-date row in the line table.
    uses_line_level=False → amounts are already on the header panels;
                            also promote Check Date to Finalized Date if present.
    uses_remittance_billed → Billed Amount and Paid Amount from remittance viewer
                            claim-details (single tab open).

    When ``is_bundle_claim``, amounts come only from the VisitDate line — no header
    fallback.

    Returns (matching_row, row_index, billed_str, paid_str).
    """
    if payer_config.get('uses_remittance_billed') and page is not None:
        prefetch_denial = payer_config.get('uses_remittance', False)
        remit_details = extract_remittance_claim_details(
            page, claim_data.get('Claim ID'),
            prefetch_denial=prefetch_denial, payer_config=payer_config,
        )
        if prefetch_denial:
            claim_data['_remit_denial'] = remit_details['denial']
        billed, paid = remit_details['billed'], remit_details['paid']
        if billed not in ('', '--'):
            log_to_gui(f"    • Billed (remittance charge): {billed}\n")
        if paid not in ('', '--'):
            log_to_gui(f"    • Paid (remittance payment): {paid}\n")
        if billed in ('', '--'):
            header_billed = claim_data.get('Billed Amount', '--')
            if header_billed not in ('', '--'):
                billed = header_billed
                log_to_gui(
                    "    ℹ️ Remittance billed not found; using header billed amount\n",
                    "info",
                )
        if paid in ('', '--'):
            header_paid = claim_data.get('Paid Amount', '--')
            if header_paid not in ('', '--'):
                paid = header_paid
                log_to_gui(
                    "    ℹ️ Remittance paid not found; using header paid amount\n",
                    "info",
                )
        claim_data['Billed Amount'] = billed
        claim_data['Paid Amount']   = paid
        return None, -1, billed, paid

    if payer_config['uses_line_level']:
        visit_date = safe_field(row_data, 'VisitDate')
        row, idx, billed, paid = find_line_by_visit_date(
            iframe, visit_date, payer_config=payer_config,
        )

        if is_bundle_claim:
            if row is None:
                billed, paid = '--', '--'
                log_to_gui(
                    f"    ⚠️ Bundle claim: no line-level row for VisitDate "
                    f"{normalize_date_range(visit_date)}\n",
                    "error",
                )
        elif row is None or (billed in ('', '--') and paid in ('', '--')):
            header_billed = claim_data.get('Billed Amount', '--')
            header_paid = claim_data.get('Paid Amount', '--')
            if header_billed not in ('', '--') or header_paid not in ('', '--'):
                billed, paid = header_billed, header_paid
                row, idx = None, -1
                log_to_gui(
                    "    ℹ️ Line-level match not found; using header billed/paid amounts\n",
                    "info",
                )

        claim_data['Billed Amount'] = billed
        claim_data['Paid Amount']   = paid
        return row, idx, billed, paid
    else:
        billed = claim_data.get('Billed Amount', '--')
        paid   = claim_data.get('Paid Amount',   '--')
        if not payer_config.get('uses_results_finalized_date'):
            if claim_data.get('Check Date') and claim_data['Check Date'] != '--':
                claim_data['Finalized Date'] = claim_data['Check Date']
        return None, -1, billed, paid


def derive_claim_status(claim_data, billed, paid, payer_config):
    """Override Claim Status where the portal-reported status is insufficient.

    Rules applied in priority order:
      1. uses_remittance payer + paid == $0           → Denied
         uses_line_level payer + paid == $0 + billed > $0 → Denied
      2. Any payer: paid > $0 but paid < billed       → Partially Paid
      3. Any payer: paid > billed                     → Paid (Overpayment)
    """
    billed_val = parse_amount(billed)
    paid_val   = parse_amount(paid)

    if paid_val is not None and paid_val == 0:
        if payer_config['uses_remittance']:
            claim_data['Claim Status'] = 'Denied'
            log_to_gui("    ℹ️ Paid $0 — marked as Denied\n", "info")
        elif (payer_config.get('uses_line_level')
                and billed_val is not None and billed_val > 0):
            claim_data['Claim Status'] = 'Denied'
            log_to_gui("    ℹ️ Paid $0 — marked as Denied\n", "info")

    if (billed_val is not None and paid_val is not None
            and billed_val > 0 and paid_val > 0 and paid_val < billed_val):
        claim_data['Claim Status'] = 'Partially Paid'
        log_to_gui("    ℹ️ Partial payment — marked as Partially Paid\n", "info")

    if billed_val is not None and paid_val is not None and paid_val > billed_val:
        claim_data['Claim Status'] = PAID_OVERPAYMENT_STATUS
        log_to_gui("    ℹ️ Paid exceeds billed — marked as Paid (Overpayment)\n", "info")


def enrich_claim_with_denial(page, iframe, claim_data, billed, paid,
                              matching_row, row_idx, payer_config, row_data=None,
                              portal_claim_status=None, is_bundle_claim=False):
    """Determine and set the Denial Reason field based on payer config.

    Dispatches to extract_denial_reason_villagecaremax() for remittance payers,
    and to extract_denial_codes_inline() for all others (when relevant).
    """
    if payer_config['uses_remittance']:
        prefetched_denial = claim_data.pop('_remit_denial', None)

        if not _has_check_number(claim_data):
            log_to_gui("    ℹ️ No check number — skipping remittance denial lookup\n", "info")
            claim_data['Denial Reason'] = '--'
            return

        status_for_skip = (portal_claim_status or claim_data.get('Claim Status', '')).strip()
        if status_for_skip.upper() == 'FINALIZED':
            log_to_gui("    ℹ️ Finalized — skipping remittance denial lookup\n", "info")
            claim_data['Denial Reason'] = '--'
            return

        billed_val = parse_amount(billed)
        paid_val   = parse_amount(paid)
        skip_denial = False
        # Fully paid or overpaid remittance claims do not need denial lookup.
        if billed_val is not None and paid_val is not None and billed_val > 0:
            if paid_val == billed_val:
                log_to_gui("    ℹ️ Fully paid — skipping remittance denial lookup\n", "info")
                skip_denial = True
            elif paid_val > billed_val:
                log_to_gui("    ℹ️ Overpayment — skipping remittance denial lookup\n", "info")
                skip_denial = True

        if skip_denial:
            claim_data['Denial Reason'] = '--'
        elif prefetched_denial is not None:
            claim_data['Denial Reason'] = prefetched_denial
        else:
            claim_data['Denial Reason'] = extract_denial_reason_villagecaremax(
                page, claim_data['Claim ID']
            )
    elif payer_config['uses_line_level']:
        claim_status = (claim_data.get('Claim Status') or '').strip().upper()
        if payer_config.get('uses_remittance_billed') and 'DENIED' not in claim_status:
            claim_data['Denial Reason'] = '--'
        elif is_bundle_claim:
            if matching_row is not None and should_extract_denial(
                claim_data['Claim Status'], billed, paid, matching_row
            ):
                claim_data['Denial Reason'] = extract_denial_codes_inline(
                    iframe, matching_row, row_idx, payer_config,
                )
            else:
                claim_data['Denial Reason'] = '--'
        else:
            visit_date = (
                coerce_visit_date(safe_field(row_data, 'VisitDate'))
                if row_data is not None else ''
            )
            claim_data['Denial Reason'] = extract_line_level_denial_reason(
                iframe, claim_data['Claim Status'], payer_config,
                visit_date=visit_date,
            )
    elif should_extract_denial(claim_data['Claim Status'], billed, paid, matching_row):
        claim_data['Denial Reason'] = extract_denial_codes_inline(
            iframe, matching_row, row_idx, payer_config,
        )
    else:
        claim_data['Denial Reason'] = '--'


# ============================================================================
# SECTION 15 — SINGLE-CLAIM PROCESSING PIPELINE
# ============================================================================

def process_one_claim(page, iframe, row_data, claim_idx, total, payer_config,
                      results_finalized_date=None, is_bundle_claim=False):
    """Extract all data fields for the claim currently open in the detail view.

    Returns a filled result dict, or empty_claim_result('Error') on failure.
    Returns None if the user has requested a stop.
    """
    if not state.is_running:
        return None

    try:
        log_to_gui(f"  → Processing claim {claim_idx + 1}/{total}...\n", "info")
        claim_data = extract_claim_header(iframe, payer_config)

        # Pending claims may still show claim number and billed/paid on the detail page
        if claim_data['Claim Status'].upper() == 'PENDING':
            log_to_gui(
                "    ℹ️ PENDING — capturing claim ID and amounts when available\n",
                "info",
            )
            enrich_claim_with_amounts(
                iframe, claim_data, row_data, payer_config,
                is_bundle_claim=is_bundle_claim, page=page,
            )
            claim_data['Denial Reason'] = '--'
            if results_finalized_date:
                claim_data['Finalized Date'] = results_finalized_date
            elif payer_config.get('uses_results_finalized_date'):
                check_date = (claim_data.get('Check Date') or '').strip()
                if check_date and check_date != '--':
                    claim_data['Finalized Date'] = check_date
            log_to_gui(f"    • Billed:   {claim_data['Billed Amount']}\n")
            log_to_gui(f"    • Paid:     {claim_data['Paid Amount']}\n")
            log_to_gui(f"    ✓ Claim {claim_idx + 1} complete (pending)\n", "success")
            return claim_data

        matching_row, row_idx, billed, paid = enrich_claim_with_amounts(
            iframe, claim_data, row_data, payer_config,
            is_bundle_claim=is_bundle_claim, page=page,
        )
        portal_claim_status = claim_data['Claim Status']
        derive_claim_status(claim_data, billed, paid, payer_config)
        enrich_claim_with_denial(
            page, iframe, claim_data, billed, paid,
            matching_row, row_idx, payer_config, row_data=row_data,
            portal_claim_status=portal_claim_status,
            is_bundle_claim=is_bundle_claim,
        )

        if results_finalized_date:
            claim_data['Finalized Date'] = results_finalized_date
        elif payer_config.get('uses_results_finalized_date'):
            check_date = (claim_data.get('Check Date') or '').strip()
            if check_date and check_date != '--':
                claim_data['Finalized Date'] = check_date
                log_to_gui(
                    f"    ℹ️ Results Finalized Date missing; using Check Date: {check_date}\n",
                    "info",
                )

        log_to_gui(f"    ✓ Claim {claim_idx + 1} complete\n", "success")
        return claim_data

    except Exception as e:
        log_to_gui(f"    ❌ Claim error: {e}\n", "error")
        return empty_claim_result('Error')


def process_all_claims_for_invoice(page, iframe, df, row_data, invoice_number,
                                   row_indices, payer_config, match_mode='invoice'):
    """Open and extract data for every matched claim results row.

    ``row_indices`` are 0-based positions in ``#claimsTable`` after sort.
    ``match_mode`` is ``'invoice'`` or ``'service_dates'`` (bundle fallback).
    After each claim (except the last), navigates back to the results list
    via the breadcrumb link inside the iframe.

    Row selection uses results-row Finalized Date identity so that after
    back-navigation re-sorts the table, we open the next *unprocessed* claim
    instead of re-opening ``row_indices[claim_idx]`` (which can point at a
    claim already extracted when sort order flips).
    Returns a list of result dicts.
    """
    results = []
    total = len(row_indices)
    is_bundle_claim = match_mode == 'service_dates'
    processed_row_keys = set()

    def _row_key_at(idx):
        row = _claim_row_locator_at_index(iframe, idx)
        finalized = _peek_results_row_finalized_date(iframe, row)
        return finalized or f'idx:{idx}', finalized, row

    for claim_idx in range(total):
        if not state.is_running:
            break

        try:
            if not state.is_running:
                break

            results_finalized_date = None
            result_row = None
            row_key = None
            for idx in row_indices:
                key, finalized, row = _row_key_at(idx)
                if key in processed_row_keys:
                    continue
                results_finalized_date = finalized
                result_row = row
                row_key = key
                break

            if result_row is None:
                log_to_gui(
                    "  ⚠️ No unprocessed matching claim rows remain\n",
                    "error",
                )
                break

            if payer_config.get('uses_results_finalized_date'):
                log_to_gui("  → Reading Finalized Date from results row...\n", "info")
                if results_finalized_date:
                    log_to_gui(
                        f"    • Results Finalized Date: {results_finalized_date}\n",
                        "info",
                    )
                else:
                    results_finalized_date = read_finalized_date_from_results_row(
                        iframe, result_row
                    )

            result_row.click()
            detail_sel = get_selector_profile(payer_config).get(
                'claim_detail_ready', '[data-testid="testClaim NumberPanel"]',
            )
            iframe.locator(detail_sel).wait_for(state='visible', timeout=15000)

            claim_data = process_one_claim(
                page, iframe, row_data, claim_idx, total, payer_config,
                results_finalized_date=(
                    results_finalized_date
                    if payer_config.get('uses_results_finalized_date')
                    else None
                ),
                is_bundle_claim=is_bundle_claim,
            )
            processed_row_keys.add(row_key)
            if claim_data:
                detail_finalized = (claim_data.get('Finalized Date') or '').strip()
                if detail_finalized and detail_finalized != '--':
                    processed_row_keys.add(detail_finalized)
                results.append(claim_data)

            if claim_idx < total - 1:
                iframe, row_indices = _navigate_back_to_results(
                    page, iframe, df, invoice_number,
                    visit_date=safe_field(row_data, 'VisitDate'),
                    payer_config=payer_config,
                    row_data=row_data,
                )
                if iframe is None:
                    break

        except Exception as e:
            log_to_gui(f"  ❌ Claim iteration error: {e}\n", "error")
            results.append(empty_claim_result('Error'))
            iframe = _attempt_recovery(page, row_data, payer_config)
            if iframe is None:
                break
            sort_claim_results_by_finalized_date(iframe)
            row_indices, recovered_mode = resolve_matching_claim_indices(
                iframe, df, invoice_number,
                visit_date=safe_field(row_data, 'VisitDate'),
                payer_config=payer_config,
            )
            if not row_indices:
                break
            is_bundle_claim = recovered_mode == 'service_dates'

    return results


def _navigate_back_to_results(page, iframe, df, invoice_number,
                              visit_date=None, payer_config=None,
                              row_data=None):
    """Click the Results breadcrumb and restore ascending sort order.

    Re-resolves matching row indices (invoice or bundle service dates) after
    sort. Callers must pick the next *unprocessed* claim by row identity
    (Finalized Date), not by a fixed ``claim_idx`` into the re-sorted list.
    Returns (iframe, row_indices) on success, or (None, None) on failure.
    """
    try:
        log_to_gui("  → Going back to results...\n")

        try:
            iframe = _return_to_results_view(page, payer_config)
        except Exception as nav_error:
            if row_data is None:
                raise nav_error
            log_to_gui(
                f"  ℹ️ Back to results via breadcrumb failed ({nav_error}) — "
                "re-running search\n",
                "info",
            )
            page, iframe = _rerun_claim_search(page, row_data, payer_config)

        sort_claim_results_by_finalized_date(iframe)
        row_indices, _match_mode = resolve_matching_claim_indices(
            iframe, df, invoice_number,
            visit_date=visit_date, payer_config=payer_config,
        )
        if not row_indices:
            raise Exception("Could not re-locate matching claims after back navigation")

        log_to_gui("  ✓ Back to results\n", "success")
        return iframe, row_indices
    except Exception as e:
        log_to_gui(f"  ❌ Back-navigation failed: {e}\n", "error")
        return None, None


def _attempt_recovery(page, row_data, payer_config=None):
    """Re-open claim search and re-run the search after a claim processing error.

    Returns a fresh iframe if successful, or None if recovery also fails.
    """
    try:
        log_to_gui("  → Attempting recovery navigation...\n", "error")
        page = _reload_and_navigate(page, payer_config)
        iframe = retry_with_backoff(
            "Recovery page ready",
            lambda: wait_for_page_ready(page, payer_config=payer_config),
            attempts=2,
        )
        fill_search_form(page, iframe, row_data, payer_config)
        submit_search_and_wait(iframe)
        return iframe
    except Exception as e:
        log_to_gui(f"  ❌ Recovery failed: {e}\n", "error")
        return None


def format_multi_claim_results(results):
    """Merge multiple claim result dicts into a single dict with numbered values.

    When one invoice has several matching claims, all values are numbered
    and joined with newlines so the output CSV stays one row per invoice.
    """
    if not results:
        return empty_claim_result('No data')
    merged_fields = [f for f in OUTPUT_COLUMNS if f != 'Last Action Taken']
    if len(results) == 1:
        return {field: results[0].get(field, '--') for field in merged_fields}
    return {
        field: '\n'.join(f"{i + 1}. {r.get(field, '--')}" for i, r in enumerate(results))
        for field in merged_fields
    }


# ============================================================================
# SECTION 16 — ROW PROCESSING  (one CSV row = one invoice)
# ============================================================================

def _reset_search_for_next_row(page, payer_config):
    """Reload the payer claim-search page so the next row starts on a clean form."""
    try:
        log_to_gui("  → Resetting search page for next row...\n")
        page = _reload_and_navigate(page, payer_config)
        retry_with_backoff(
            "Reset search page ready",
            lambda: wait_for_page_ready(page, payer_config=payer_config),
            attempts=3,
        )
        return page
    except Exception as e:
        log_to_gui(f"  ⚠️ Search page reset failed: {e}\n", "error")
        return page


def _reset_or_return_to_results(page, payer_config, *, skip_row_reset=False):
    """After a row completes, reload search form or stay on results for grouping."""
    if skip_row_reset:
        try:
            page = get_live_page(page)
            if not _results_table_is_visible(page, payer_config):
                _return_to_results_view(page, payer_config)
            else:
                wait_for_claim_results_ready(page, payer_config=payer_config, quiet=True)
            sort_claim_results_by_finalized_date(
                page.frame_locator(get_selector_profile(payer_config)['iframe']),
            )
        except Exception as e:
            log_to_gui(
                f"  ⚠️ Could not return to results ({e}) — full reset\n",
                "error",
            )
            return _reset_search_for_next_row(page, payer_config)
        return page
    return _reset_search_for_next_row(page, payer_config)


def _finish_row_without_claims(
    page,
    df,
    row_index,
    output_folder,
    payer_config,
    *,
    claim_status,
    automation_status='Done',
    last_error='',
    success_log=None,
    reload_after=False,
    skip_row_reset=False,
):
    """Record a non-extraction row outcome, reset search, then save progress."""
    write_row_result(df, row_index, empty_claim_result(claim_status))
    safe_mark_row_status(df, row_index, automation_status, last_error)
    if success_log:
        log_to_gui(success_log + "\n", "success")
    if reload_after:
        page = _reset_or_return_to_results(
            page, payer_config, skip_row_reset=skip_row_reset,
        )
    save_progress_results(df, output_folder)
    return page


def process_one_row(page, df, row_index, row_data, output_folder, payer_config,
                    *, skip_initial_navigate=False, skip_search=False,
                    skip_row_reset=False):
    """Full pipeline for one CSV row: navigate → search → extract → save.

    Returns ``(row_ok, page)`` where ``row_ok`` is True when the row was
    handled (including intentional skips).  ``page`` is reset to the claim
    search form after successful extraction so the next row starts clean.
    """
    if not state.is_running:
        return False, page

    try:
        log_to_gui(f"\n🔄 Row {row_index + 1}\n", "info")
        invoice_number = safe_invoice_number(row_data, 'InvoiceNumber')
        visit_date     = safe_field(row_data, 'VisitDate')

        if not invoice_number:
            log_to_gui(f"  ⚠️ Row {row_index + 1}: InvoiceNumber is empty — skipping\n", "error")
            write_row_result(df, row_index, empty_claim_result('Missing InvoiceNumber'))
            safe_mark_row_status(df, row_index, 'Error', 'Missing InvoiceNumber')
            save_progress_results(df, output_folder, force=True)
            return False, page

        if not row_has_valid_dob(row_data, safe_field):
            dob_raw = safe_field(row_data, 'DOB')
            log_to_gui(
                f"  ⚠️ Row {row_index + 1}: DOB missing or invalid "
                f"({dob_raw!r}) — skipping\n",
                "error",
            )
            page = _finish_row_without_claims(
                page, df, row_index, output_folder, payer_config,
                claim_status=DOB_NOT_FOUND_STATUS,
                automation_status='Skipped',
                success_log=f"  ✓ Row {row_index + 1} skipped ({DOB_NOT_FOUND_STATUS})",
                reload_after=False,
            )
            return True, page

        log_to_gui(f"  → Invoice: {invoice_number}  |  Visit: {visit_date}\n")

        initial_start, initial_end = resolve_service_search_dates(row_data, safe_field)
        bundle_retry_attempted = False

        if skip_search:
            log_to_gui("  ℹ️ Reusing search results (grouped row)\n", "info")
            try:
                page = get_live_page(page)
                if not _results_table_is_visible(page, payer_config):
                    iframe = _return_to_results_view(page, payer_config)
                else:
                    iframe = wait_for_claim_results_ready(
                        page, payer_config=payer_config, quiet=True,
                    )
                sort_claim_results_by_finalized_date(iframe)
            except Exception as nav_error:
                log_to_gui(
                    f"  ❌ Could not reuse search results: {nav_error}\n",
                    "error",
                )
                write_row_result(df, row_index, empty_claim_result('Search failed'))
                safe_mark_row_status(
                    df, row_index, 'Error', f'Search failed: {nav_error}',
                )
                save_progress_results(df, output_folder, force=True)
                return False, page
            search_outcome = 'results'
            bundle_retry_attempted = True
        else:
            # First row in batch: session already verified.  All later rows reset
            # to the payer search URL regardless of spreadsheet row index.
            if skip_initial_navigate:
                try:
                    iframe = retry_with_backoff(
                        "Wait for page ready",
                        lambda: wait_for_page_ready(page, payer_config=payer_config),
                        attempts=3,
                    )
                except Exception as nav_error:
                    log_to_gui(f"  ❌ Page setup failed: {nav_error}\n", "error")
                    write_row_result(df, row_index, empty_claim_result('Navigation failed'))
                    safe_mark_row_status(df, row_index, 'Error', f'Navigation failed: {nav_error}')
                    save_progress_results(df, output_folder, force=True)
                    return False, page
            else:
                page = _reload_and_navigate(page, payer_config)
                try:
                    iframe = retry_with_backoff(
                        "Wait for page ready",
                        lambda: wait_for_page_ready(page, payer_config=payer_config),
                        attempts=3,
                    )
                except Exception as nav_error:
                    log_to_gui(f"  ❌ Page setup failed: {nav_error}\n", "error")
                    write_row_result(df, row_index, empty_claim_result('Navigation failed'))
                    safe_mark_row_status(df, row_index, 'Error', f'Navigation failed: {nav_error}')
                    save_progress_results(df, output_folder, force=True)
                    return False, page

            if not fill_search_form(page, iframe, row_data, payer_config):
                outcome, page, iframe = _handle_form_fill_failure(
                    page, iframe, df, row_index, row_data, output_folder, payer_config,
                )
                if outcome == 'failed':
                    return False, page
                if outcome == 'row_done':
                    return True, page

            search_outcome = submit_search_and_wait(iframe)
            if search_outcome == 'not_found':
                retry_result = _try_bundle_search_date_fallback(
                    page, iframe, df, invoice_number, row_data, payer_config,
                    initial_start, initial_end,
                    already_attempted=bundle_retry_attempted,
                )
                if retry_result:
                    bundle_retry_attempted = True
                    search_outcome, page, iframe = retry_result
            if search_outcome == 'results':
                sort_claim_results_by_finalized_date(iframe)

        if search_outcome == 'invalid_member_id':
            page = _finish_row_without_claims(
                page, df, row_index, output_folder, payer_config,
                claim_status=INVALID_MEMBER_ID_STATUS,
                success_log=f"  ✓ Row {row_index + 1} complete ({INVALID_MEMBER_ID_STATUS})",
            )
            return True, page

        if search_outcome == 'not_found':
            page = _finish_row_without_claims(
                page, df, row_index, output_folder, payer_config,
                claim_status=CLAIM_NOT_FOUND_STATUS,
                success_log=f"  ✓ Row {row_index + 1} complete (claim not found)",
            )
            return True, page

        if search_outcome == 'payer_error':
            status = _resolve_payer_error_status()
            page = _finish_row_without_claims(
                page, df, row_index, output_folder, payer_config,
                claim_status=status,
                success_log=f"  ✓ Row {row_index + 1} complete ({status})",
            )
            return True, page

        if search_outcome != 'results':
            page = _finish_row_without_claims(
                page, df, row_index, output_folder, payer_config,
                claim_status='No claims found. Advised to search manually',
                last_error='Search failed',
                success_log=f"  ✓ Row {row_index + 1} complete (search failed — moving on)",
            )
            return True, page

        row_indices, match_mode = resolve_matching_claim_indices(
            iframe, df, invoice_number,
            visit_date=visit_date, payer_config=payer_config,
        )
        if not row_indices:
            retry_result = _try_bundle_search_date_fallback(
                page, iframe, df, invoice_number, row_data, payer_config,
                initial_start, initial_end,
                already_attempted=bundle_retry_attempted,
            )
            if retry_result:
                bundle_retry_attempted = True
                retry_outcome, page, iframe = retry_result
            else:
                retry_outcome = None
            if retry_outcome == 'results':
                row_indices, match_mode = resolve_matching_claim_indices(
                    iframe, df, invoice_number,
                    visit_date=visit_date, payer_config=payer_config,
                )
            elif retry_outcome:
                search_outcome = retry_outcome
                if search_outcome == 'not_found':
                    page = _finish_row_without_claims(
                        page, df, row_index, output_folder, payer_config,
                        claim_status=CLAIM_NOT_FOUND_STATUS,
                        success_log=f"  ✓ Row {row_index + 1} complete (claim not found)",
                    )
                    return True, page
                if search_outcome == 'invalid_member_id':
                    page = _finish_row_without_claims(
                        page, df, row_index, output_folder, payer_config,
                        claim_status=INVALID_MEMBER_ID_STATUS,
                        success_log=f"  ✓ Row {row_index + 1} complete ({INVALID_MEMBER_ID_STATUS})",
                    )
                    return True, page
                if search_outcome == 'payer_error':
                    status = _resolve_payer_error_status()
                    page = _finish_row_without_claims(
                        page, df, row_index, output_folder, payer_config,
                        claim_status=status,
                        success_log=f"  ✓ Row {row_index + 1} complete ({status})",
                    )
                    return True, page
                page = _finish_row_without_claims(
                    page, df, row_index, output_folder, payer_config,
                    claim_status='No claims found. Advised to search manually',
                    last_error='Search failed',
                    success_log=f"  ✓ Row {row_index + 1} complete (search failed — moving on)",
                )
                return True, page

        if not row_indices:
            page = _finish_row_without_claims(
                page, df, row_index, output_folder, payer_config,
                claim_status='Claim not found. Search manually',
                automation_status='Error',
                last_error='Claim not found',
                success_log=f"  ✓ Row {row_index + 1} complete (claim not found in results)",
            )
            return True, page

        if match_mode == 'service_dates':
            log_to_gui("  ℹ️ Bundle claim match (service dates)\n", "info")

        all_results = process_all_claims_for_invoice(
            page, iframe, df, row_data, invoice_number, row_indices, payer_config,
            match_mode=match_mode,
        )
        write_row_result(df, row_index, format_multi_claim_results(all_results))
        safe_mark_row_status(df, row_index, 'Done', '')
        log_to_gui(f"  ✓ Row {row_index + 1} complete\n", "success")
        save_progress_results(df, output_folder)
        page = _reset_or_return_to_results(
            page, payer_config, skip_row_reset=skip_row_reset,
        )
        return True, page

    except Exception as e:
        log_to_gui(f"  ❌ Row error: {e}\n", "error")
        log_to_gui(traceback.format_exc(), "error")
        write_row_result(df, row_index, empty_claim_result('Critical error'))
        safe_mark_row_status(df, row_index, 'Error', str(e))
        save_progress_results(df, output_folder, force=True)
        capture_row_failure_artifact(
            page, output_folder, row_index,
            {
                'error': str(e),
                'payer': state.current_payer_name or '',
                'invoice': safe_invoice_number(row_data, 'InvoiceNumber', ''),
            },
        )
        try:
            page = _reset_or_return_to_results(
                page, payer_config, skip_row_reset=False,
            )
        except Exception:
            pass
        return False, page


def process_search_group(page, df, group_indices, output_folder, payer_config,
                         *, skip_initial_navigate=False):
    """Run one claim search for ``group_indices`` then extract each row.

    Returns ``(all_ok, page, row_results)`` where ``row_results`` maps each
    index to True/False (handled ok vs error).
    """
    if not group_indices:
        return True, page, {}

    row_results = {}
    first_idx = group_indices[0]
    last_idx = group_indices[-1]

    log_to_gui(
        f"\n📋 Search group: {len(group_indices)} row(s) "
        f"({first_idx + 1}–{last_idx + 1}) sharing one search\n",
        "info",
    )

    row_ok, page = process_one_row(
        page, df, first_idx, df.iloc[first_idx], output_folder, payer_config,
        skip_initial_navigate=skip_initial_navigate,
        skip_search=False,
        skip_row_reset=len(group_indices) > 1,
    )
    row_results[first_idx] = row_ok

    if not row_ok:
        for idx in group_indices[1:]:
            solo_ok, page = process_one_row(
                page, df, idx, df.iloc[idx], output_folder, payer_config,
                skip_initial_navigate=False,
            )
            row_results[idx] = solo_ok
        return all(row_results.values()), page, row_results

    for idx in group_indices[1:]:
        if not state.is_running:
            break
        row_ok, page = process_one_row(
            page, df, idx, df.iloc[idx], output_folder, payer_config,
            skip_initial_navigate=True,
            skip_search=True,
            skip_row_reset=(idx != last_idx),
        )
        if not row_ok:
            log_to_gui(
                f"  ↻ Row {idx + 1} failed in group — retrying with solo search\n",
                "info",
            )
            row_ok, page = process_one_row(
                page, df, idx, df.iloc[idx], output_folder, payer_config,
                skip_initial_navigate=False,
            )
        row_results[idx] = row_ok

    all_ok = all(row_results.get(i, False) for i in group_indices)
    return all_ok, page, row_results


def _goto_navigation_home(page):
    """Unload the embedded Availity app so the next payer load remounts the iframe."""
    page.goto(
        AVAILITY_NAVIGATION_HOME,
        wait_until='domcontentloaded',
        timeout=45000,
    )
    page.wait_for_load_state('domcontentloaded', timeout=15000)
    return page


def _reload_and_navigate(page, payer_config):
    """Reset to a clean payer claim-search form via two-hop shell navigation.

    A same-URL ``goto`` from claim detail often leaves ``iframe#newBodyFrame`` on
    the detail route.  Unloading the navigation shell first forces a fresh app load.
    Breadcrumb shortcuts are not used here — MUI service-date pickers keep stale
    values after a prior search and reject in-place date changes.
    """
    log_to_gui("  → Navigating to search page...\n")
    page = retry_with_backoff(
        "Acquire live page",
        lambda: get_live_page(page),
        attempts=3,
    )
    retry_with_backoff(
        "Unload Availity app shell",
        lambda: _goto_navigation_home(page),
        attempts=3,
    )
    page = navigate_to_payer_page(page, payer_config)
    log_to_gui("  ✓ Navigation complete\n", "success")
    return page


# ============================================================================
# SECTION 17 — BATCH PROCESSING
# ============================================================================

def open_edge_and_connect(payer_name, excel_path):
    """Phase 1: load Excel, then launch Edge with CDP and normalize tabs.

    Runs in a background thread (spawned by the GUI). Excel is read before any
    browser work. The user signs in to Availity manually in Edge; only after
    they click Start Automation do we navigate to claim search and process rows.

    Note on threading: ``sync_playwright()`` objects are bound to the thread
    that called ``start()`` and cannot be used (or even ``stop()``-ed) from
    another thread. So we attach Playwright here only long enough to bring
    the right tab forward, then release it on this same thread. Phase 2
    re-connects fresh on its own worker thread.
    """
    state.worker_busy = True
    user_stopped = False
    try:
        payer_config = PAYER_CONFIG.get(payer_name)
        if not payer_config:
            log_to_gui(f"❌ Unknown payer '{payer_name}'\n", "error")
            _set_ui_state(UI_STATE_IDLE)
            return

        log_to_gui("📂 Loading Excel file...\n", "info")
        df, stats = load_and_prepare_excel(excel_path)
        if df is None:
            state.is_running = False
            _set_ui_state(UI_STATE_IDLE)
            return

        state.loaded_df = df
        state.loaded_excel_path = excel_path
        state.excel_row_stats = stats
        update_excel_row_stats(stats)
        log_to_gui(f"✓ Loaded {stats['total']} row(s)\n", "success")
        log_to_gui(
            f"   • Total rows read: {stats['total']}\n"
            f"   • Rows with Recheck = yes: {stats['recheck_yes']}\n"
            f"   • Rows with Recheck empty (new): {stats['recheck_empty']}\n",
            "info",
        )
        if not state.is_running:
            return

        log_to_gui("🌐 Starting Edge browser...\n", "info")
        if not ensure_edge_with_cdp():
            state.is_running = False
            _set_ui_state(UI_STATE_IDLE)
            return
        if not state.is_running:
            return

        log_to_gui("🌐 Connecting to browser via CDP...\n", "info")
        page = setup_browser()
        # Release Playwright on the same thread that started it. Phase 2 will
        # re-attach on its own worker thread before it does any work.
        ensure_playwright_disconnected()

        if not state.is_running:
            return

        state.selected_payer_config = payer_config
        state.selected_payer_name = payer_name
        state.current_payer_name = payer_name
        log_to_gui("✓ Browser connected\n", "success")
        log_to_gui(
            "✓ Edge ready. Please sign in to Availity in Edge, then click "
            "Start Automation.\n",
            "success",
        )
        _set_ui_state(UI_STATE_READY)
    except Exception as e:
        log_to_gui(f"❌ Could not open Edge: {e}\n", "error")
        log_to_gui(traceback.format_exc(), "error")
        state.is_running = False
        _set_ui_state(UI_STATE_IDLE)
    finally:
        user_stopped = not state.is_running
        state.worker_busy = False
        ensure_playwright_disconnected()
        if user_stopped:
            close_managed_edge_if_owned()
            reset_ui_state(clear_all=False)




def run_batch(batch_size, excel_path, output_folder):
    """Phase 2: process rows from the Excel loaded in Phase 1.

    Expects ``state.selected_payer_config`` and ``state.loaded_df`` to be set by
    ``open_edge_and_connect`` and Edge to still be running. Re-attaches
    Playwright on this worker thread because Playwright sync objects can't
    cross threads.
    """
    df = None
    should_close_browser = False
    page = None
    payer_config = state.selected_payer_config
    user_stopped = False

    state.worker_busy = True
    state.stop_save_requested = False
    state.rows_since_last_save = 0
    state._large_autofit_notice_logged = False
    try:
        if payer_config is None:
            log_to_gui(
                "❌ Browser is not open. Click Open Edge & Login first.\n",
                "error",
            )
            return

        payer_name = state.selected_payer_name or ''
        if not payer_name:
            for name, cfg in PAYER_CONFIG.items():
                if cfg is payer_config:
                    payer_name = name
                    break
        if not payer_name:
            payer_name = payer_config.get('display_name') or payer_config.get('name') or ''
        state.current_payer_name = payer_name

        if state.loaded_df is None or state.loaded_excel_path != excel_path:
            log_to_gui("📂 Reloading Excel file...\n", "info")
            df, stats = load_and_prepare_excel(excel_path)
            if df is None:
                return
            state.loaded_df = df
            state.loaded_excel_path = excel_path
            state.excel_row_stats = stats
            update_excel_row_stats(stats)
        else:
            df = state.loaded_df.copy()

        stats = state.excel_row_stats or compute_recheck_stats(df)
        update_excel_row_stats(stats)
        log_to_gui(
            f"✓ Using {stats.get('total', len(df))} row(s) — "
            f"Recheck yes: {stats.get('recheck_yes', 0)}, "
            f"new (empty Recheck): {stats.get('recheck_empty', 0)}\n",
            "success",
        )

        log_to_gui("▶ Starting automation...\n", "info")
        log_to_gui("  → Attaching Playwright to Edge...\n", "info")
        page = setup_browser()

        log_to_gui("  → Verifying Availity session...\n", "info")
        ok, page = _verify_payer_search_session(page, payer_config, quiet=True)
        if not ok:
            log_to_gui(
                "❌ Could not open the payer claim search. Sign in to Availity "
                "in Edge, then click Start Automation again.\n",
                "error",
            )
            return
        log_to_gui("✓ Session verified — claim search is open.\n", "success")

        df = queue_recheck_yes_rows(df)
        df = queue_error_rows_for_retry(df)

        # Expose for emergency save from the GUI thread (Stop button).
        state.current_batch_df = df
        state.current_batch_output_folder = output_folder
        _remit_amount_paths.clear()
        hydrate_remit_amount_paths(output_folder, payer_name, _remit_amount_paths)
        start_save_worker()
        pending_indices = [idx for idx in range(len(df)) if row_should_process(df, idx)]
        target_indices = pending_indices[:batch_size]
        work_queue = list(target_indices)
        initial_queued = len(work_queue)
        rows_to_process = len(work_queue)
        batch_error_indices = set()

        if rows_to_process == 0:
            log_to_gui(
                "✓ No rows to process (all Done/Skipped with no prior errors).\n",
                "success",
            )
            reset_automation_progress("Nothing to process — all rows done", 100)
            should_close_browser = True
            _run_finalize_save(df, output_folder)
            return

        log_to_gui(
            f"\n📦 Processing {initial_queued} pending row(s)"
            f"{' for ' + payer_name if payer_name else ''} (resume-aware)...\n",
            "info"
        )
        log_to_gui("-" * 60 + "\n")

        update_automation_progress(
            0,
            rows_to_process,
            f"Starting batch — {rows_to_process} claim(s) queued",
        )

        attempted_count = 0
        error_counts = {}
        search_form_ready = True
        circuit_breaker = BatchCircuitBreaker()
        search_groups = build_consecutive_search_groups(work_queue, df, payer_config)
        group_pos = 0

        def _record_batch_row(idx, row_ok, page_ref):
            nonlocal page, search_form_ready
            if row_ok:
                search_form_ready = True
                batch_error_indices.discard(idx)
                circuit_breaker.record_row_success()
            else:
                search_form_ready = False
                batch_error_indices.add(idx)
                failure_stage = classify_row_failure(
                    df.at[idx, 'LastError'],
                    df.at[idx, 'Claim Status'],
                )
                circuit_breaker.record_transient_failure(failure_stage)
                capture_row_failure_artifact(
                    page_ref, output_folder, idx,
                    {
                        'payer': payer_name,
                        'invoice': safe_invoice_number(df.iloc[idx], 'InvoiceNumber', ''),
                        'last_error': str(df.at[idx, 'LastError']),
                        'failure_stage': failure_stage,
                    },
                )
            if not row_ok and str(df.at[idx, 'AutomationStatus']).strip().lower() != 'error':
                safe_mark_row_status(df, idx, 'Error', 'Unknown row processing error')
                save_progress_results(df, output_folder, force=True)
            _record_error(df, idx, error_counts)

        while group_pos < len(search_groups):
            group = search_groups[group_pos]
            idx = group[0]
            if not state.is_running:
                log_to_gui("⚠ Stopped by user\n", "error")
                break

            if circuit_breaker.tripped:
                log_to_gui(
                    f"⚠ Circuit breaker tripped: {circuit_breaker.trip_reason}\n",
                    "error",
                )
                break

            active_page = _find_logged_in_page(page)
            if active_page is not None:
                page = active_page
            ready, reason = validate_navigation_page_ready(page, timeout_ms=2500)
            if not ready:
                circuit_breaker.record_session_failure()
                if circuit_breaker.tripped:
                    log_to_gui(
                        f"⚠ Circuit breaker tripped: {circuit_breaker.trip_reason}\n",
                        "error",
                    )
                    break
                log_to_gui(
                    "⚠ Session appears logged out. Awaiting user re-login...\n",
                    "error",
                )
                log_to_gui(f"    Session check reason: {reason}\n", "error")

                state.resume_event.clear()
                state.resume_decision = "cancel"
                if state.root is not None:
                    state.root.after(0, lambda r=reason: _prompt_user_relogin(r))
                state.resume_event.wait()

                if state.resume_decision != "resume" or not state.is_running:
                    log_to_gui(
                        "⚠ User cancelled after session expiry.\n",
                        "error",
                    )
                    break

                log_to_gui("  → Re-verifying Availity session...\n", "info")
                ok, page = _verify_payer_search_session(page, payer_config, quiet=True)
                if not ok:
                    log_to_gui(
                        "❌ Session still not usable; stopping batch.\n",
                        "error",
                    )
                    break
                log_to_gui("✓ Session re-verified — resuming.\n", "success")
                circuit_breaker.record_session_recovery()
                search_form_ready = True

                prior_errors = sorted(i for i in batch_error_indices if i != idx)
                if prior_errors:
                    log_to_gui(
                        f"  ↻ Resuming: retrying {len(prior_errors)} prior error row(s), "
                        f"then row {idx + 1} and continuing…\n",
                        "info",
                    )
                tail = work_queue[group_pos:]
                flat_tail = [i for g in search_groups[group_pos:] for i in g]
                work_queue = _dedupe_preserve_order(sorted(batch_error_indices) + flat_tail)
                rows_to_process = len(work_queue)
                search_groups = build_consecutive_search_groups(work_queue, df, payer_config)
                group_pos = 0
                continue

            if len(group) >= 2:
                progress_label = (
                    f"Processing search group rows {group[0] + 1}–{group[-1] + 1} "
                    f"({len(group)} rows, one search)"
                )
            else:
                is_prior_error_retry = idx in batch_error_indices
                progress_label = (
                    f"Retrying row {idx + 1} (prior error)"
                    if is_prior_error_retry
                    else f"Processing row {idx + 1}"
                )
            invoice_number = safe_invoice_number(df.iloc[idx], 'InvoiceNumber', '')
            update_automation_progress(
                attempted_count,
                rows_to_process,
                progress_label
                + (f" — Invoice {invoice_number}" if invoice_number else ""),
            )

            if len(group) >= 2:
                for row_idx in group:
                    safe_mark_row_status(df, row_idx, 'InProgress', '')
                _, page, row_results = process_search_group(
                    page, df, group, output_folder, payer_config,
                    skip_initial_navigate=search_form_ready,
                )
                for row_idx, row_ok in row_results.items():
                    attempted_count += 1
                    _record_batch_row(row_idx, row_ok, page)
                    update_automation_progress(
                        attempted_count,
                        rows_to_process,
                        f"Completed {attempted_count} of {rows_to_process} attempts — "
                        f"row {row_idx + 1}"
                        + (
                            f" — Invoice {safe_invoice_number(df.iloc[row_idx], 'InvoiceNumber', '')}"
                            if safe_invoice_number(df.iloc[row_idx], 'InvoiceNumber', '')
                            else ""
                        ),
                    )
                    if circuit_breaker.tripped:
                        break
            else:
                safe_mark_row_status(df, idx, 'InProgress', '')
                row_ok, page = process_one_row(
                    page, df, idx, df.iloc[idx], output_folder, payer_config,
                    skip_initial_navigate=search_form_ready,
                )
                attempted_count += 1
                _record_batch_row(idx, row_ok, page)
                update_automation_progress(
                    attempted_count,
                    rows_to_process,
                    f"Completed {attempted_count} of {rows_to_process} attempts — row {idx + 1}"
                    + (f" — Invoice {invoice_number}" if invoice_number else ""),
                )

            if circuit_breaker.tripped:
                log_to_gui(
                    f"⚠ Circuit breaker tripped: {circuit_breaker.trip_reason}\n",
                    "error",
                )
                break

            group_pos += 1

        log_to_gui("-" * 60 + "\n")
        if not state.stop_save_requested:
            _run_finalize_save(df, output_folder)
        else:
            log_to_gui("  ℹ️ Final save handled by Stop — skipping duplicate.\n", "info")

        batch_complete = group_pos >= len(search_groups) and state.is_running
        if batch_complete:
            log_to_gui(
                f"✓ Complete! {attempted_count} attempt(s), "
                f"{initial_queued} initially queued\n",
                "success",
            )
            finish_automation_progress(attempted_count, rows_to_process, complete=True)
            should_close_browser = True
        else:
            log_to_gui(
                f"⚠ Stopped. {attempted_count}/{rows_to_process} attempt(s) saved\n",
                "error",
            )
            finish_automation_progress(attempted_count, rows_to_process, stopped=True)

        if error_counts:
            log_to_gui("\n📊 Error summary:\n", "error")
            for reason, count in sorted(error_counts.items(), key=lambda x: -x[1]):
                log_to_gui(f"   • {reason}: {count} row(s)\n", "error")

    except Exception as e:
        log_to_gui(f"❌ Critical batch error: {e}\n", "error")
        log_to_gui(traceback.format_exc(), "error")
        if df is not None and not state.stop_save_requested:
            _run_finalize_save(df, output_folder)
    finally:
        user_stopped = not state.is_running
        state.worker_busy = False
        stop_save_worker(wait=True)
        state.current_batch_df = None
        state.current_batch_output_folder = None
        state.current_payer_name = None
        state.loaded_df = None
        state.loaded_excel_path = None
        state.excel_row_stats = None
        ensure_playwright_disconnected()
        if should_close_browser or user_stopped:
            close_managed_edge_if_owned()
        reset_ui_state(clear_all=user_stopped)


def _record_error(df, row_index, error_counts):
    """Increment the counter for a known error status written into Claim Status."""
    status_val = df.at[row_index, 'Claim Status']
    if not isinstance(status_val, str):
        return
    for line in status_val.split('\n'):
        val = line.split('. ', 1)[-1].strip()
        if val in KNOWN_ERROR_STATUSES:
            error_counts[val] = error_counts.get(val, 0) + 1

